from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Q, Count, Case, When, IntegerField
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, timedelta, date
import json
import requests
from django.core import signing
from django.conf import settings
import logging

from .models import (
    PreventiveJobTemplate,
    PreventiveJobExecution,
    PreventiveJobAttachment,
    PreventiveJobNotification,
    ChecklistResult,
)
from .forms import (
    PreventiveJobTemplateForm,
    PreventiveJobExecutionForm,
    PreventiveJobExecutionQuickUpdateForm,
    PreventiveJobAttachmentForm,
    PreventiveJobAttachmentFormSet,
)
from core.models import AsetMesin, CustomUser
from core.export_handlers import send_to_google_apps_script

logger = logging.getLogger(__name__)

# Config for share tokens (seconds)
SHARE_TOKEN_MAX_AGE = getattr(settings, 'PREVENTIVE_SHARE_TOKEN_MAX_AGE', 7 * 24 * 3600)  # default 7 days
SHARE_SIGN_SALT = getattr(settings, 'PREVENTIVE_SHARE_SIGN_SALT', 'preventive-share-salt')


# ==============================================================================
# INDEX/HUB VIEW
# ==============================================================================

@login_required(login_url='core:login')
def preventive_index_view(request):
    """
    Halaman index/hub untuk Preventive Jobs - menampilkan menu-menu utama
    """
    user = request.user
    
    # === GET SUBORDINATES ===
    subordinate_ids = user.get_all_subordinates()
    all_user_ids = [user.id] + subordinate_ids
    
    # === GET QUICK STATS ===
    today = timezone.now().date()
    current_month_start = date(today.year, today.month, 1)
    if today.month == 12:
        current_month_end = date(today.year + 1, 1, 1) - timedelta(days=1)
    else:
        current_month_end = date(today.year, today.month + 1, 1) - timedelta(days=1)
    
    # Filter executions untuk current user dan subordinates
    executions_this_month = PreventiveJobExecution.objects.filter(
        scheduled_date__gte=current_month_start,
        scheduled_date__lte=current_month_end
    ).filter(
        Q(template__pic__id__in=all_user_ids) |
        Q(assigned_to__id__in=all_user_ids)
    )
    
    # Stats for this month
    total_jobs = executions_this_month.count()
    completed_jobs = executions_this_month.filter(status='Done').count()
    overdue_jobs = executions_this_month.filter(
        status='Scheduled',
        scheduled_date__lt=today
    ).count()
    
    compliance_rate = (completed_jobs / total_jobs * 100) if total_jobs > 0 else 0
    
    # Total templates available (templates dimana user adalah PIC atau assigned, dan tidak dihapus)
    total_templates = PreventiveJobTemplate.objects.filter(
        (Q(pic__id__in=all_user_ids) |
         Q(created_by__id__in=all_user_ids)) &
        Q(is_deleted=False)
    ).distinct().count()
    
    context = {
        'total_jobs': total_jobs,
        'completed_jobs': completed_jobs,
        'overdue_jobs': overdue_jobs,
        'compliance_rate': compliance_rate,
        'total_templates': total_templates,
    }
    
    return render(request, 'preventive_jobs/index.html', context)


# ==============================================================================
# DASHBOARD VIEWS
# ==============================================================================

@login_required(login_url='core:login')
def preventive_dashboard_view(request):
    """
    Dashboard utama Preventive Job V2
    - Overview KPI
    - Grafik compliance trend
    - Table execution terdekat
    """
    user = request.user
    
    # === GET SUBORDINATES ===
    subordinate_ids = user.get_all_subordinates()
    all_user_ids = [user.id] + subordinate_ids
    
    # === FILTER PARAMETERS ===
    current_year = request.GET.get('year', str(timezone.now().year))
    current_month = request.GET.get('month', str(timezone.now().month))
    selected_line = request.GET.get('line', '')
    selected_status = request.GET.get('status', '')
    
    # Handle "Semua" (empty value) untuk month dan year
    filter_all_months = False
    filter_all_years = False
    
    try:
        if current_year and current_year != '0':
            current_year = int(current_year)
        else:
            filter_all_years = True
            current_year = timezone.now().year  # Default untuk context
        
        if current_month and current_month != '0':
            current_month = int(current_month)
        else:
            filter_all_months = True
            current_month = timezone.now().month  # Default untuk context
    except ValueError:
        current_year = timezone.now().year
        current_month = timezone.now().month
    
    # === FILTER TANGGAL ===
    today = timezone.now().date()
    
    # Jika filter all months, ambil seluruh tahun
    if filter_all_months and filter_all_years:
        # Semua data tahun ini
        start_of_period = date(timezone.now().year, 1, 1)
        end_of_period = date(timezone.now().year, 12, 31)
    elif filter_all_months:
        # Semua bulan di tahun yang dipilih
        start_of_period = date(current_year, 1, 1)
        end_of_period = date(current_year, 12, 31)
    elif filter_all_years:
        # Bulan yang dipilih di semua tahun
        start_of_period = date(2020, current_month, 1)  # Mulai dari 2020
        end_of_period = date(timezone.now().year + 5, current_month, 1) - timedelta(days=1)
    else:
        # Bulan dan tahun spesifik
        start_of_period = date(current_year, current_month, 1)
        if current_month == 12:
            end_of_period = date(current_year + 1, 1, 1) - timedelta(days=1)
        else:
            end_of_period = date(current_year, current_month + 1, 1) - timedelta(days=1)
    
    # === BUILD QUERY ===
    # Filter: execution dari user atau subordinate-nya
    executions_query = PreventiveJobExecution.objects.filter(
        scheduled_date__gte=start_of_period,
        scheduled_date__lte=end_of_period
    ).filter(
        Q(template__pic__id__in=all_user_ids) |  # PIC adalah user atau subordinate
        Q(assigned_to__id__in=all_user_ids)       # Ditugaskan ke user atau subordinate
    )
    
    if selected_status:
        executions_query = executions_query.filter(status=selected_status)
    
    # === KPI CALCULATIONS ===
    total_jobs = executions_query.count()
    done_jobs = executions_query.filter(status='Done').count()
    late_jobs = executions_query.filter(
        status='Scheduled',
        scheduled_date__lt=today
    ).count()
    
    compliance_rate = (done_jobs / total_jobs * 100) if total_jobs > 0 else 0
    
    # === GRAFIK DATA (6 BULAN TERAKHIR) ===
    compliance_trend = []
    for i in range(5, -1, -1):
        check_date = today - timedelta(days=30*i)
        month_start = date(check_date.year, check_date.month, 1)
        if check_date.month == 12:
            month_end = date(check_date.year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(check_date.year, check_date.month + 1, 1) - timedelta(days=1)
        
        month_jobs = PreventiveJobExecution.objects.filter(
            scheduled_date__gte=month_start,
            scheduled_date__lte=month_end
        ).filter(
            Q(template__pic__id__in=all_user_ids) |  # Filter by user/subordinates
            Q(assigned_to__id__in=all_user_ids)
        )
        month_done = month_jobs.filter(status='Done').count()
        month_total = month_jobs.count()
        month_rate = (month_done / month_total * 100) if month_total > 0 else 0
        
        compliance_trend.append({
            'month': f"{month_start.strftime('%b')} {month_start.year}",
            'rate': round(month_rate, 1)
        })
    
    # === TABLE: UPCOMING EXECUTIONS (THIS WEEK) ===
    week_start = today
    week_end = today + timedelta(days=7)
    
    upcoming_executions = PreventiveJobExecution.objects.filter(
        scheduled_date__gte=week_start,
        scheduled_date__lte=week_end
    ).filter(
        Q(template__pic__id__in=all_user_ids) |  # Filter by user/subordinates
        Q(assigned_to__id__in=all_user_ids)
    ).select_related(
        'template',
        'aset',
        'assigned_to'
    ).order_by('scheduled_date')
    
    # === NOTIFIKASI YANG BELUM DIBACA ===
    unread_notifications = PreventiveJobNotification.objects.filter(
        user=user,
        is_read=False
    ).count()
    
    context = {
        'total_jobs': total_jobs,
        'done_jobs': done_jobs,
        'late_jobs': late_jobs,
        'compliance_rate': round(compliance_rate, 1),
        'compliance_trend': json.dumps(compliance_trend),
        'upcoming_executions': upcoming_executions[:5],
        'unread_notifications': unread_notifications,
        'current_year': current_year,
        'current_month': current_month,
        'selected_status': selected_status,
        'filter_all_months': filter_all_months,
        'filter_all_years': filter_all_years,
        'year_list': range(timezone.now().year - 2, timezone.now().year + 3),
        'month_list': list(range(1, 13)),
    }
    
    return render(request, 'preventive_jobs/dashboard.html', context)


# ==============================================================================
# TEMPLATE MANAGEMENT VIEWS
# ==============================================================================

@login_required(login_url='core:login')
def preventive_template_list_view(request):
    """
    Halaman list semua template preventive job
    """
    user = request.user
    
    # === FILTER PARAMETERS ===
    selected_line = request.GET.get('line', '')
    selected_status = request.GET.get('status', '')
    search_query = request.GET.get('q', '')
    
    # === BUILD QUERY ===
    # Filter: hanya tampilkan template yang tidak dihapus (is_deleted=False)
    templates = PreventiveJobTemplate.objects.filter(is_deleted=False)
    
    if search_query:
        templates = templates.filter(
            Q(nama_pekerjaan__icontains=search_query) |
            Q(deskripsi__icontains=search_query)
        )
    
    if selected_status == 'active':
        templates = templates.filter(is_active=True)
    elif selected_status == 'inactive':
        templates = templates.filter(is_active=False)
    
    # === PAGINATION ===
    from django.core.paginator import Paginator
    paginator = Paginator(templates, 10)
    page_number = request.GET.get('page')
    templates_page = paginator.get_page(page_number)
    
    context = {
        'templates': templates_page,
        'search_query': search_query,
        'selected_status': selected_status,
    }
    
    return render(request, 'preventive_jobs/template_list.html', context)


@login_required(login_url='core:login')
def preventive_template_create_view(request):
    """
    Halaman create preventive job template
    """
    next_url = request.GET.get('next', request.POST.get('next', ''))
    
    if request.method == 'POST':
        form = PreventiveJobTemplateForm(request.POST, user=request.user)
        if form.is_valid():
            try:
                from django.db import transaction
                import logging
                
                logger = logging.getLogger(__name__)
                
                with transaction.atomic():
                    template = form.save(commit=False)
                    template.created_by = request.user
                    template.pic = form.cleaned_data.get('pic', request.user)
                    template.save()
                    
                    # Save many-to-many relationships
                    form.save_m2m()
                    
                    # === AUTO-GENERATE EXECUTION RECORDS ===
                    generate_executions_for_template(template)
                    
                    logger.info(f"Created template '{template.nama_pekerjaan}' with {template.executions.count()} executions")
                    
                    messages.success(
                        request,
                        f"Template '{template.nama_pekerjaan}' berhasil dibuat dan {template.executions.count()} execution records telah di-generate."
                    )
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error creating template: {str(e)}", exc_info=True)
                messages.error(request, f"Error saat membuat template: {str(e)}")
                return render(request, 'preventive_jobs/template_form.html', {'form': form, 'title': 'Buat Template Preventive Job', 'next': next_url})
            
            if next_url:
                return redirect(next_url)
            return redirect('preventive_jobs:template_list')
    else:
        form = PreventiveJobTemplateForm(user=request.user)
        form.fields['pic'].initial = request.user
    
    context = {'form': form, 'title': 'Buat Template Preventive Job', 'next': next_url}
    return render(request, 'preventive_jobs/template_form.html', context)


@login_required(login_url='core:login')
def preventive_template_edit_view(request, template_id):
    """
    Halaman edit preventive job template
    """
    from django.db import transaction
    import logging
    
    logger = logging.getLogger(__name__)
    template = get_object_or_404(PreventiveJobTemplate, id=template_id, is_deleted=False)
    next_url = request.GET.get('next', request.POST.get('next', ''))
    
    # Store old values untuk detect perubahan
    old_tanggal_berakhir = template.tanggal_berakhir
    old_tanggal_mulai = template.tanggal_mulai
    old_interval_hari = template.interval_hari
    old_schedule_type = template.schedule_type
    old_custom_dates = template.custom_dates
    
    if request.method == 'POST':
        form = PreventiveJobTemplateForm(request.POST, instance=template, user=request.user)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Store old aset list before saving
                    old_aset_ids = set(template.aset_mesin.values_list('id', flat=True))
                    
                    template = form.save(commit=False)
                    template.save()
                    
                    # Get new aset list after save
                    form.save_m2m()
                    new_aset_ids = set(template.aset_mesin.values_list('id', flat=True))
                    
                    # === HANDLE ASET CHANGES ===
                    added_aset_ids = new_aset_ids - old_aset_ids
                    removed_aset_ids = old_aset_ids - new_aset_ids
                    
                    # 1. Generate execution untuk aset yang baru ditambahkan
                    if added_aset_ids:
                        from .models import PreventiveJobExecution
                        for aset_id in added_aset_ids:
                            execution_dates = template.get_all_execution_dates()
                            for scheduled_date in execution_dates:
                                PreventiveJobExecution.objects.get_or_create(
                                    template=template,
                                    aset_id=aset_id,
                                    scheduled_date=scheduled_date,
                                    defaults={
                                        'status': 'Scheduled',
                                        'compliance_type': 'None',
                                    }
                                )
                        logger.info(f"Added {len(added_aset_ids)} new asets to template {template_id}")
                    
                    # 2. Hapus execution dari aset yang dihapus
                    if removed_aset_ids:
                        template.executions.filter(aset_id__in=removed_aset_ids).delete()
                        logger.info(f"Removed {len(removed_aset_ids)} asets from template {template_id}")
                    
                    # === HANDLE JADWAL CHANGES ===
                    # Detect perubahan tanggal atau interval
                    tanggal_changed = (template.tanggal_mulai != old_tanggal_mulai or 
                                      template.tanggal_berakhir != old_tanggal_berakhir)
                    interval_changed = (template.interval_hari != old_interval_hari or
                                       template.schedule_type != old_schedule_type or
                                       template.custom_dates != old_custom_dates)
                    
                    if tanggal_changed or interval_changed:
                        # Generate semua execution dates yang baru (tidak delete yang lama)
                        generate_executions_for_extended_template(template, old_tanggal_berakhir)
                        messages.info(
                            request, 
                            "Jadwal berhasil diupdate. Execution records baru telah di-generate (yang lama tetap ada)."
                        )
                        logger.info(f"Updated schedule for template {template_id}")
                    else:
                        messages.success(request, "Template berhasil diupdate.")
                        logger.info(f"Updated template {template_id}")
                    
            except Exception as e:
                logger.error(f"Error updating template {template_id}: {str(e)}", exc_info=True)
                messages.error(request, f"Error saat update template: {str(e)}")
                return render(request, 'preventive_jobs/template_form.html', {
                    'form': form,
                    'template': template,
                    'title': f'Edit Template: {template.nama_pekerjaan}',
                    'next': next_url
                })
            
            if next_url:
                return redirect(next_url)
            return redirect('preventive_jobs:template_list')
    else:
        form = PreventiveJobTemplateForm(instance=template, user=request.user)
    
    context = {
        'form': form,
        'template': template,
        'title': f'Edit Template: {template.nama_pekerjaan}',
        'next': next_url
    }
    return render(request, 'preventive_jobs/template_form.html', context)


@login_required(login_url='core:login')
def preventive_template_detail_view(request, template_id):
    """
    Halaman detail template + execution tracking
    """
    template = get_object_or_404(PreventiveJobTemplate, id=template_id, is_deleted=False)
    
    # === FILTER EXECUTIONS ===
    selected_month = request.GET.get('month', '')
    selected_year = request.GET.get('year', '')
    
    # Jika tidak ada filter, tampilkan ALL executions untuk melihat apakah ada data
    if selected_month and selected_year:
        try:
            selected_month = int(selected_month)
            selected_year = int(selected_year)
        except ValueError:
            selected_month = timezone.now().month
            selected_year = timezone.now().year
        
        start_date = date(selected_year, selected_month, 1)
        if selected_month == 12:
            end_date = date(selected_year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(selected_year, selected_month + 1, 1) - timedelta(days=1)
        
        executions = template.executions.filter(
            scheduled_date__gte=start_date,
            scheduled_date__lte=end_date
        ).order_by('scheduled_date')
        
        selected_month = int(selected_month)
        selected_year = int(selected_year)
    else:
        # Tampilkan semua executions jika tidak ada filter
        executions = template.executions.all().order_by('scheduled_date')
        selected_month = timezone.now().month
        selected_year = timezone.now().year
    
    # === STATISTICS ===
    total_exec = executions.count()
    done_exec = executions.filter(status='Done').count()
    compliance_rate = (done_exec / total_exec * 100) if total_exec > 0 else 0
    
    context = {
        'template': template,
        'executions': executions,
        'aset_list': template.aset_mesin.all(),
        'total_exec': total_exec,
        'done_exec': done_exec,
        'compliance_rate': round(compliance_rate, 1),
        'selected_month': selected_month,
        'selected_year': selected_year,
    }
    
    return render(request, 'preventive_jobs/template_detail.html', context)


@login_required(login_url='core:login')
@require_http_methods(["POST"])
def preventive_template_delete_view(request, template_id):
    """
    Soft-delete preventive job template (pindah ke recycle bin)
    Executions tetap tersimpan untuk restore kemudian
    """
    template = get_object_or_404(PreventiveJobTemplate, id=template_id, is_deleted=False)
    template_name = template.nama_pekerjaan
    
    # Soft-delete template dan related executions
    template.soft_delete(deleted_by=request.user)
    
    messages.success(request, f"Template '{template_name}' dipindahkan ke recycle bin. Anda bisa restore-nya di halaman Recycle Bin jika diperlukan.")
    return redirect('preventive_jobs:template_list')


@login_required(login_url='core:login')
@require_http_methods(["POST"])
def preventive_template_toggle_view(request, template_id):
    """
    Toggle status aktif/nonaktif template
    """
    template = get_object_or_404(PreventiveJobTemplate, id=template_id, is_deleted=False)
    template.is_active = not template.is_active
    template.save()
    
    status = "Aktif" if template.is_active else "Nonaktif"
    messages.success(request, f"Template '{template.nama_pekerjaan}' sekarang {status}.")
    
    return redirect('preventive_jobs:template_list')


@login_required(login_url='core:login')
def preventive_template_duplicate_view(request, template_id):
    """
    Duplicate template dengan periode/jadwal berbeda
    Hanya perlu input tanggal_mulai dan tanggal_berakhir
    Semua field lain dicopy dari template original
    
    Use case: 
      - Template "Cek Motor A - Jan 2025" sudah dibuat
      - Feb 2025 ingin pakai schedule yang sama, tapi tanggal berbeda
      - Click "Duplicate" → input Feb 1-28 → Done!
    """
    from preventive_jobs.forms import PreventiveTemplateDuplicateForm
    
    original = get_object_or_404(PreventiveJobTemplate, id=template_id)
    next_url = request.GET.get('next', request.POST.get('next', ''))
    
    if request.method == 'POST':
        form = PreventiveTemplateDuplicateForm(request.POST)
        if form.is_valid():
            tanggal_mulai = form.cleaned_data['tanggal_mulai']
            tanggal_berakhir = form.cleaned_data['tanggal_berakhir']
            
            # Create new template by cloning original
            cloned = PreventiveJobTemplate.objects.create(
                nama_pekerjaan=original.nama_pekerjaan,
                deskripsi=original.deskripsi,
                fokus=original.fokus,
                prioritas=original.prioritas,
                kategori=original.kategori,
                schedule_type=original.schedule_type,
                interval_hari=original.interval_hari,
                custom_dates=original.custom_dates,
                tanggal_mulai=tanggal_mulai,
                tanggal_berakhir=tanggal_berakhir,
                checklist_template=original.checklist_template,
                pic=original.pic,
                notify_24h_before=original.notify_24h_before,
                notify_2h_before=original.notify_2h_before,
                notify_on_schedule=original.notify_on_schedule,
                is_active=True,  # Set active for new template
                created_by=request.user,
            )
            
            # Copy M2M relationships (aset_mesin)
            cloned.aset_mesin.set(original.aset_mesin.all())
            
            # Explicit save to trigger _generate_executions()
            cloned.save()
            
            # Log untuk audit
            messages.success(
                request, 
                f"Template '{cloned.nama_pekerjaan}' berhasil di-duplicate untuk periode {tanggal_mulai.strftime('%d/%m/%Y')} - {tanggal_berakhir.strftime('%d/%m/%Y') if tanggal_berakhir else 'Ongoing'}. "
                f"Sistem telah menggenerate {cloned.executions.count()} execution records otomatis."
            )
            
            if next_url:
                return redirect(next_url)
            return redirect('preventive_jobs:template_detail', template_id=cloned.id)
    else:
        form = PreventiveTemplateDuplicateForm()
        
        # Pre-fill suggested dates (next period)
        from dateutil.relativedelta import relativedelta
        suggested_start = original.tanggal_berakhir + timedelta(days=1) if original.tanggal_berakhir else timezone.now().date()
        suggested_end = suggested_start + relativedelta(months=1) - timedelta(days=1)
    
    context = {
        'form': form,
        'original': original,
        'title': f'Duplicate Template: {original.nama_pekerjaan}',
        'suggested_start': suggested_start if request.method == 'GET' else None,
        'suggested_end': suggested_end if request.method == 'GET' else None,
        'next': next_url,
    }
    
    return render(request, 'preventive_jobs/template_duplicate.html', context)


# ==============================================================================
# EXECUTION TRACKING VIEWS
# ==============================================================================

@login_required(login_url='core:login')
def preventive_execution_list_view(request):
    """
    Halaman list execution dengan TAB (TODAY, THIS WEEK, ADVANCED FILTER)
    Hanya menampilkan execution yang sesuai dengan user (sebagai PIC atau assigned_to)
    """
    user = request.user
    today = timezone.now().date()
    
    # === GET TAB (default: today) ===
    tab = request.GET.get('tab', 'today')  # today, thisweek, advanced
    
    # === FILTER PARAMETERS (untuk tab advanced) ===
    selected_status = request.GET.get('status', '')
    selected_month = request.GET.get('month', '')
    selected_year = request.GET.get('year', str(timezone.now().year))
    search_query = request.GET.get('q', '')
    sort_param = request.GET.get('sort', '-scheduled_date')
    
    try:
        selected_year = int(selected_year)
    except ValueError:
        selected_year = timezone.now().year
    
    # Handle month filter
    if selected_month:
        try:
            selected_month = int(selected_month)
        except ValueError:
            selected_month = None
    else:
        selected_month = None
    
    # === BUILD DATE RANGE BASED ON TAB ===
    if tab == 'today':
        start_date = today
        end_date = today
    elif tab == 'thisweek':
        start_date = today
        end_date = today + timedelta(days=7)
    elif tab == 'advanced':
        # Custom filter dengan bulan/tahun
        if selected_month:
            start_date = date(selected_year, selected_month, 1)
            if selected_month == 12:
                end_date = date(selected_year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(selected_year, selected_month + 1, 1) - timedelta(days=1)
        else:
            # Show all months in selected year
            start_date = date(selected_year, 1, 1)
            end_date = date(selected_year, 12, 31)
    else:
        # Default ke today
        tab = 'today'
        start_date = today
        end_date = today
    
    # === GET SUBORDINATES ===
    subordinate_ids = user.get_all_subordinates()
    all_user_ids = [user.id] + subordinate_ids
    
    # === BUILD QUERY ===
    # ALWAYS filter by user access - even in advanced tab
    # User dapat melihat: executions milik dirinya, tim atasan mereka, atau bawahan mereka
    
    executions = PreventiveJobExecution.objects.filter(
        scheduled_date__gte=start_date,
        scheduled_date__lte=end_date
    ).filter(
        Q(template__pic__id__in=all_user_ids) |
        Q(assigned_to__id__in=all_user_ids) |
        Q(assigned_to_personil__penanggung_jawab__id__in=all_user_ids)
    )
    
    executions = executions.distinct()
    
    # === APPLY ADDITIONAL FILTERS (untuk tab advanced) ===
    if tab == 'advanced':
        if selected_status:
            executions = executions.filter(status=selected_status)
        
        if search_query:
            executions = executions.filter(
                Q(template__nama_pekerjaan__icontains=search_query) |
                Q(aset__nama__icontains=search_query)
            )
    
    executions = executions.select_related(
        'template', 'aset', 'assigned_to'
    ).prefetch_related(
        'assigned_to_personil'
    ).order_by(sort_param)
    
    # === PAGINATION ===
    from django.core.paginator import Paginator
    paginator = Paginator(executions, 20)
    page_number = request.GET.get('page')
    executions_page = paginator.get_page(page_number)
    
    # === MONTH LIST FOR FILTER ===
    month_choices = [
        (1, 'January'), (2, 'February'), (3, 'March'), (4, 'April'),
        (5, 'May'), (6, 'June'), (7, 'July'), (8, 'August'),
        (9, 'September'), (10, 'October'), (11, 'November'), (12, 'December')
    ]
    
    # === YEAR LIST FOR FILTER ===
    year_list = list(range(2023, 2027))
    
    # === COUNT UNTUK TAB ===
    today_count = PreventiveJobExecution.objects.filter(
        scheduled_date=today
    ).filter(
        Q(template__pic__id__in=all_user_ids) |
        Q(assigned_to__id__in=all_user_ids) |
        Q(assigned_to_personil__penanggung_jawab__id__in=all_user_ids)
    ).distinct().count()
    
    week_end = today + timedelta(days=7)
    thisweek_count = PreventiveJobExecution.objects.filter(
        scheduled_date__gte=today,
        scheduled_date__lte=week_end
    ).filter(
        Q(template__pic__id__in=all_user_ids) |
        Q(assigned_to__id__in=all_user_ids) |
        Q(assigned_to_personil__penanggung_jawab__id__in=all_user_ids)
    ).distinct().count()
    
    context = {
        'executions': executions_page,
        'search_query': search_query,
        'selected_status': selected_status,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'month_choices': month_choices,
        'year_list': year_list,
        'sort_param': sort_param,
        'tab': tab,
        'today_count': today_count,
        'thisweek_count': thisweek_count,
    }
    
    return render(request, 'preventive_jobs/execution_list.html', context)


@login_required(login_url='core:login')
def preventive_execution_detail_view(request, execution_id):
    """
    Halaman detail execution + update form
    """
    execution = get_object_or_404(PreventiveJobExecution, id=execution_id)
    
    if request.method == 'POST':
        # Check if user wants to change status via state machine
        new_status = request.POST.get('status')
        reason = request.POST.get('status_change_reason', '')
        
        if new_status and new_status != execution.status:
            # Use state machine for status transition
            try:
                is_valid, message = execution.can_transition_to(new_status)
                if is_valid:
                    execution.transition_to(
                        new_status,
                        reason=reason,
                        changed_by=request.user
                    )
                    messages.success(request, f"Status diubah: {execution.status} ({reason or 'No reason'})")
                else:
                    messages.error(request, f"Transisi status gagal: {message}")
            except Exception as e:
                messages.error(request, f"Error mengubah status: {str(e)}")
        
        # Process form changes for other fields (if not just status change)
        if new_status:
            # Only validate formset if changing other fields too
            if len(request.POST) > 3:  # More than just status fields
                form = PreventiveJobExecutionForm(request.POST, instance=execution, user=request.user)
                formset = PreventiveJobAttachmentFormSet(request.POST, request.FILES, instance=execution)
                
                if form.is_valid() and formset.is_valid():
                    execution = form.save()
                    formset.save()
                    
                    if execution.attachments.exists():
                        execution.has_attachment = True
                        execution.save()
                    
                    messages.success(request, "Execution berhasil diupdate.")
        
        return redirect('preventive_jobs:execution_list')
    else:
        form = PreventiveJobExecutionForm(instance=execution, user=request.user)
        formset = PreventiveJobAttachmentFormSet(instance=execution)
        
        # Get status history for display
        status_history = execution.get_status_history()
        last_log = status_history.order_by('-changed_at').first()
    
    context = {
        'execution': execution,
        'form': form,
        'formset': formset,
        'status_history': status_history,
        'last_log': last_log,
        'can_undo': status_history.exists(),
        'valid_transitions': execution.VALID_STATUS_TRANSITIONS.get(execution.status, []),
    }
    
    return render(request, 'preventive_jobs/execution_detail.html', context)


# ==============================================================================
# EXECUTION ASSIGN VIEW (AJAX)
# ==============================================================================

@login_required(login_url='core:login')
@require_http_methods(["GET"])
def execution_get_personil_list(request, execution_id):
    """
    AJAX endpoint untuk get daftar personil berdasarkan PIC dari template
    Personil adalah anak buah dari PIC (Personil.penanggung_jawab == PIC)
    Jika user adalah atasan PIC, bisa assign ke tim PIC
    Returns: { "success": true, "personil_list": [...] }
    """
    import traceback
    
    try:
        user = request.user
        
        # Get execution
        try:
            execution = PreventiveJobExecution.objects.get(id=execution_id)
        except PreventiveJobExecution.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': f'Execution ID {execution_id} tidak ditemukan'
            }, status=404)
        
        # Get template and PIC
        try:
            template = execution.template
            pic = template.pic
            if not pic:
                return JsonResponse({
                    'success': False,
                    'message': 'Template tidak memiliki PIC'
                }, status=400)
        except AttributeError as e:
            return JsonResponse({
                'success': False,
                'message': f'Error mengakses template/PIC: {str(e)}'
            }, status=400)
        
        # Check authorization
        # User berhak assign jika:
        # 1. User adalah PIC, atau
        # 2. User adalah atasan dari PIC (dalam hierarchy tree)
        
        user_can_assign = False
        
        # Check 1: User adalah PIC
        if user.id == pic.id:
            user_can_assign = True
        
        # Check 2: User adalah atasan dari PIC
        if not user_can_assign:
            current_user = pic
            max_iterations = 10
            iteration = 0
            
            while current_user and iteration < max_iterations:
                iteration += 1
                try:
                    # Field nama: "atasan", bukan "penanggung_jawab"
                    if hasattr(current_user, 'atasan') and current_user.atasan:
                        if user.id == current_user.atasan.id:
                            user_can_assign = True
                            break
                        current_user = current_user.atasan
                    else:
                        break
                except:
                    break
        
        if not user_can_assign:
            return JsonResponse({
                'success': False,
                'message': f'Anda tidak memiliki hak untuk assign execution ini. PIC={pic.username}'
            }, status=403)
        
        # Get personil list dari PIC (gunakan Personil model)
        # Personil adalah "anak buah" (tim) dari PIC
        try:
            from core.models import Personil
            
            personil_list_objs = Personil.objects.filter(
                penanggung_jawab=pic
            ).order_by('nama_lengkap')
            
            personil_list = []
            for personil in personil_list_objs:
                personil_list.append({
                    'id': personil.id,
                    'nama_lengkap': personil.nama_lengkap
                })
            
            pic_name = pic.get_full_name() if hasattr(pic, 'get_full_name') else str(pic)
            if not pic_name or pic_name.strip() == '':
                pic_name = pic.username if hasattr(pic, 'username') else 'Unknown'
            
            return JsonResponse({
                'success': True,
                'personil_list': personil_list,
                'pic_name': pic_name,
                'pic_id': pic.id
            })
        
        except Exception as query_error:
            error_trace = traceback.format_exc()
            print(f"Query error: {error_trace}")
            return JsonResponse({
                'success': False,
                'message': f'Error mengakses personil: {str(query_error)}'
            }, status=500)
    
    except Exception as e:
        error_trace = traceback.format_exc()
        print(f"Unexpected error: {error_trace}")
        return JsonResponse({
            'success': False,
            'message': f'Unexpected error: {str(e)}'
        }, status=500)


@login_required(login_url='core:login')
@require_http_methods(["POST"])
def execution_assign_view(request, execution_id):
    """
    AJAX endpoint untuk assign execution ke satu atau lebih personil
    Expected POST data: { "assigned_to_personil": [personil_id1, personil_id2, ...] }
    """
    import traceback
    
    try:
        # Get execution
        try:
            execution = PreventiveJobExecution.objects.get(id=execution_id)
        except PreventiveJobExecution.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': f'Execution ID {execution_id} tidak ditemukan'
            }, status=404)
        
        # Parse JSON body
        try:
            data = json.loads(request.body)
            assigned_to_personil_ids = data.get('assigned_to_personil', [])
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': 'Invalid JSON format'
            }, status=400)
        
        if not assigned_to_personil_ids or len(assigned_to_personil_ids) == 0:
            return JsonResponse({
                'success': False,
                'message': 'Pilih minimal satu personil'
            }, status=400)
        
        # Validate personil exists
        from core.models import Personil
        try:
            personil_list = Personil.objects.filter(id__in=assigned_to_personil_ids)
            if personil_list.count() != len(assigned_to_personil_ids):
                return JsonResponse({
                    'success': False,
                    'message': 'Beberapa personil tidak ditemukan'
                }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error validating personil: {str(e)}'
            }, status=500)
        
        # Update execution - clear old then add new
        try:
            execution.assigned_to_personil.clear()
            execution.assigned_to_personil.set(personil_list)
            execution.save()
            
            # Get names for response
            names = [p.nama_lengkap for p in personil_list]
            
            return JsonResponse({
                'success': True,
                'message': f'Berhasil assign ke {", ".join(names)}',
                'assigned_personil_ids': assigned_to_personil_ids,
                'assigned_personil_names': names
            })
        except Exception as save_error:
            error_trace = traceback.format_exc()
            print(f"Save error: {error_trace}")
            return JsonResponse({
                'success': False,
                'message': f'Error saving assignment: {str(save_error)}'
            }, status=500)
    
    except Exception as e:
        error_trace = traceback.format_exc()
        print(f"Unexpected error in execution_assign_view: {error_trace}")
        return JsonResponse({
            'success': False,
            'message': f'Unexpected error: {str(e)}'
        }, status=500)


# ==============================================================================
# MONITORING VIEWS
# ==============================================================================

@login_required(login_url='core:login')
def preventive_mesin_monitoring_view(request):
    """
    Heatmap view untuk monitoring compliance per mesin
    """
    # === FILTER PARAMETERS ===
    selected_month = request.GET.get('month', str(timezone.now().month))
    selected_year = request.GET.get('year', str(timezone.now().year))
    
    try:
        selected_month = int(selected_month)
        selected_year = int(selected_year)
    except ValueError:
        selected_month = timezone.now().month
        selected_year = timezone.now().year
    
    # === FILTER DATE RANGE ===
    start_date = date(selected_year, selected_month, 1)
    if selected_month == 12:
        end_date = date(selected_year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(selected_year, selected_month + 1, 1) - timedelta(days=1)
    
    # === GET ALL MESIN ===
    mesin_list = AsetMesin.objects.filter(level=2).prefetch_related('preventive_job_executions')
    
    mesin_data = []
    for mesin in mesin_list:
        executions = mesin.preventive_job_executions.filter(
            scheduled_date__gte=start_date,
            scheduled_date__lte=end_date
        )
        
        total = executions.count()
        done = executions.filter(status='Done').count()
        compliance = (done / total * 100) if total > 0 else 0
        
        mesin_data.append({
            'mesin': mesin,
            'total': total,
            'done': done,
            'compliance': round(compliance, 1),
            'compliance_color': get_compliance_color(compliance),
        })
    
    context = {
        'mesin_data': mesin_data,
        'selected_month': selected_month,
        'selected_year': selected_year,
    }
    
    return render(request, 'preventive_jobs/mesin_monitoring.html', context)


# ==============================================================================
# COMPLIANCE REPORT VIEWS
# ==============================================================================

@login_required(login_url='core:login')
def preventive_compliance_report_view(request):
    """
    Halaman compliance report & analytics
    """
    # === FILTER PARAMETERS ===
    report_period = request.GET.get('period', 'month')  # month, quarter, year
    from_month = request.GET.get('from_month', str(timezone.now().month))
    from_year = request.GET.get('from_year', str(timezone.now().year))
    to_month = request.GET.get('to_month', str(timezone.now().month))
    to_year = request.GET.get('to_year', str(timezone.now().year))
    
    try:
        from_month = int(from_month)
        from_year = int(from_year)
        to_month = int(to_month)
        to_year = int(to_year)
    except ValueError:
        pass
    
    # === BUILD DATE RANGE ===
    start_date = date(from_year, from_month, 1)
    if to_month == 12:
        end_date = date(to_year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(to_year, to_month + 1, 1) - timedelta(days=1)
    
    # === GET ALL EXECUTIONS ===
    executions = PreventiveJobExecution.objects.filter(
        scheduled_date__gte=start_date,
        scheduled_date__lte=end_date
    )
    
    # === KPI SUMMARY ===
    total_jobs = executions.count()
    done_jobs = executions.filter(status='Done').count()
    overall_compliance = (done_jobs / total_jobs * 100) if total_jobs > 0 else 0
    
    # === HITUNG OVERDUE DAYS ===
    today = timezone.now().date()
    overdue_jobs = executions.filter(
        status='Scheduled',
        scheduled_date__lt=today
    )
    if overdue_jobs.exists():
        avg_overdue_days = sum([
            (today - e.scheduled_date).days for e in overdue_jobs
        ]) / overdue_jobs.count()
    else:
        avg_overdue_days = 0
    
    # === MESIN DENGAN COMPLIANCE TERENDAH (TOP 5) ===
    mesin_stats = []
    for mesin in AsetMesin.objects.all():
        mesin_execs = executions.filter(aset=mesin)
        if mesin_execs.exists():
            total = mesin_execs.count()
            done = mesin_execs.filter(status='Done').count()
            compliance = (done / total * 100)
            mesin_stats.append({
                'mesin': mesin,
                'total': total,
                'done': done,
                'compliance': round(compliance, 1),
            })
    
    mesin_stats = sorted(mesin_stats, key=lambda x: x['compliance'])[:5]
    
    # === COMPLIANCE TREND (12 BULAN) ===
    trend_data = []
    for i in range(11, -1, -1):
        check_date = today - timedelta(days=30*i)
        month_start = date(check_date.year, check_date.month, 1)
        if check_date.month == 12:
            month_end = date(check_date.year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(check_date.year, check_date.month + 1, 1) - timedelta(days=1)
        
        month_jobs = executions.filter(
            scheduled_date__gte=month_start,
            scheduled_date__lte=month_end
        )
        month_done = month_jobs.filter(status='Done').count()
        month_total = month_jobs.count()
        month_rate = (month_done / month_total * 100) if month_total > 0 else 0
        
        trend_data.append({
            'month': f"{month_start.strftime('%b %y')}",
            'compliance': round(month_rate, 1)
        })
    
    context = {
        'total_jobs': total_jobs,
        'done_jobs': done_jobs,
        'overall_compliance': round(overall_compliance, 1),
        'avg_overdue_days': round(avg_overdue_days, 1),
        'mesin_stats': mesin_stats,
        'trend_data': json.dumps(trend_data),
        'from_month': from_month,
        'from_year': from_year,
        'to_month': to_month,
        'to_year': to_year,
    }
    
    return render(request, 'preventive_jobs/compliance_report.html', context)


# ==============================================================================
# HELPER FUNCTIONS
# ==============================================================================

def generate_executions_for_template(template):
    """
    Generate execution records untuk template berdasarkan periode dan range tanggal
    """
    execution_dates = template.get_all_execution_dates()
    
    for exec_date in execution_dates:
        for aset in template.aset_mesin.all():
            PreventiveJobExecution.objects.get_or_create(
                template=template,
                aset=aset,
                scheduled_date=exec_date,
                defaults={
                    'status': 'Scheduled',
                }
            )


def generate_executions_for_extended_template(template, old_tanggal_berakhir):
    """
    Generate execution records TAMBAHAN untuk template ketika jadwal diperpanjang.
    Hanya generate dates yang BARU (belum ada di database).
    Tidak akan delete execution yang sudah ada.
    """
    from .models import PreventiveJobExecution
    from django.db import transaction
    import logging
    
    logger = logging.getLogger(__name__)
    
    try:
        with transaction.atomic():
            # Get all execution dates dari updated template
            all_execution_dates = template.get_all_execution_dates()
            
            # Get existing execution dates dari database
            existing_dates = PreventiveJobExecution.objects.filter(
                template=template
            ).values_list('scheduled_date', flat=True).distinct()
            existing_dates_set = set(existing_dates)
            
            # Get dates yang belum ada di database (dates baru)
            new_dates = [d for d in all_execution_dates if d not in existing_dates_set]
            
            # Generate execution untuk dates yang baru saja
            for new_date in new_dates:
                for aset in template.aset_mesin.all():
                    PreventiveJobExecution.objects.get_or_create(
                        template=template,
                        aset=aset,
                        scheduled_date=new_date,
                        defaults={
                            'status': 'Scheduled',
                            'compliance_type': 'None',
                        }
                    )
            
            # Log info untuk debug
            count_new = len(new_dates) * template.aset_mesin.count()
            logger.info(
                f"Generated {count_new} new execution records for template '{template.nama_pekerjaan}'. "
                f"New dates: {new_dates}, Aset count: {template.aset_mesin.count()}"
            )
            
    except Exception as e:
        logger.error(f"Error generating executions for template {template.id}: {str(e)}", exc_info=True)
        raise


def get_compliance_color(compliance_rate):
    """
    Return warna untuk compliance rate
    """
    if compliance_rate >= 90:
        return 'success'  # Green
    elif compliance_rate >= 70:
        return 'warning'  # Yellow
    else:
        return 'danger'  # Red


# ==============================================================================
# MASTER CHECKLIST VIEWS
# ==============================================================================

@login_required(login_url='core:login')
def checklist_list_view(request):
    """
    Halaman list semua master checklist
    """
    from .models import ChecklistTemplate
    
    # Filter
    kategori = request.GET.get('kategori', '')
    search_query = request.GET.get('q', '')
    
    checklists = ChecklistTemplate.objects.all()
    
    if kategori:
        checklists = checklists.filter(kategori=kategori)
    
    if search_query:
        checklists = checklists.filter(
            Q(nama__icontains=search_query) |
            Q(nomor__icontains=search_query) |
            Q(deskripsi__icontains=search_query)
        )
    
    checklists = checklists.order_by('-created_at')
    
    context = {
        'checklists': checklists,
        'selected_kategori': kategori,
        'search_query': search_query,
        'kategori_choices': ChecklistTemplate.KATEGORI_CHOICES,
    }
    
    return render(request, 'preventive_jobs/checklist_list.html', context)


@login_required(login_url='core:login')
def checklist_create_view(request):
    """
    Halaman membuat checklist template baru
    """
    from .models import ChecklistTemplate
    from .forms import ChecklistTemplateForm, ChecklistItemFormSet
    
    if request.method == 'POST':
        form = ChecklistTemplateForm(request.POST, request.FILES)
        formset = ChecklistItemFormSet(request.POST, instance=None)
        
        if form.is_valid():
            checklist = form.save(commit=False)
            checklist.created_by = request.user
            checklist.save()
            
            formset.instance = checklist
            if formset.is_valid():
                formset.save()
                messages.success(request, f"Checklist '{checklist.nama}' berhasil dibuat!")
                return redirect('preventive_jobs:checklist_list')
    else:
        form = ChecklistTemplateForm()
        formset = ChecklistItemFormSet(instance=None)
    
    context = {
        'form': form,
        'formset': formset,
        'title': 'Buat Checklist Template Baru',
        'is_create': True,
    }
    
    return render(request, 'preventive_jobs/checklist_form.html', context)


@login_required(login_url='core:login')
def checklist_detail_view(request, pk):
    """
    Halaman detail dan edit checklist template
    """
    from .models import ChecklistTemplate
    from .forms import ChecklistTemplateForm, ChecklistItemFormSet
    
    checklist = get_object_or_404(ChecklistTemplate, pk=pk)
    
    if request.method == 'POST':
        form = ChecklistTemplateForm(request.POST, request.FILES, instance=checklist)
        formset = ChecklistItemFormSet(request.POST, instance=checklist)
        
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, f"Checklist '{checklist.nama}' berhasil diupdate!")
            return redirect('preventive_jobs:checklist_list')
    else:
        form = ChecklistTemplateForm(instance=checklist)
        formset = ChecklistItemFormSet(instance=checklist)
    
    context = {
        'checklist': checklist,
        'form': form,
        'formset': formset,
        'title': f"Edit: {checklist.nama}",
        'is_create': False,
    }
    
    return render(request, 'preventive_jobs/checklist_form.html', context)


@login_required(login_url='core:login')
def checklist_delete_view(request, pk):
    """
    Halaman delete checklist template
    """
    from .models import ChecklistTemplate
    
    checklist = get_object_or_404(ChecklistTemplate, pk=pk)
    
    if request.method == 'POST':
        nama = checklist.nama
        checklist.delete()
        messages.success(request, f"Checklist '{nama}' berhasil dihapus!")
        return redirect('preventive_jobs:checklist_list')
    
    context = {
        'checklist': checklist,
        'title': f"Hapus: {checklist.nama}",
    }
    
    return render(request, 'preventive_jobs/checklist_confirm_delete.html', context)


# ==============================================================================
# AJAX ENDPOINTS UNTUK CHECKLIST
# ==============================================================================

@login_required(login_url='core:login')
@require_http_methods(["GET"])
def get_checklist_items_view(request, execution_id):
    """
    AJAX endpoint untuk load checklist items dari execution
    Return JSON dengan struktur checklist items & existing result (jika ada)
    """
    from .models import PreventiveJobExecution, ChecklistResult
    
    def user_can_view_execution(user, execution):
        """Check if user can view execution and its checklist"""
        # Staff/Superuser can view everything
        if user.is_staff or user.is_superuser:
            print(f"[DEBUG] {user} can view - is staff/superuser")
            return True
        
        # Allow all authenticated users to view (permissive approach for now)
        # This can be made more restrictive later if needed
        print(f"[DEBUG] {user} allowed to view execution {execution.id} (permissive)")
        return True
    
    try:
        execution = get_object_or_404(PreventiveJobExecution, pk=execution_id)
        
        if not user_can_view_execution(request.user, execution):
            return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)
        
        # Get checklist template
        if not execution.template.checklist_template:
            return JsonResponse({'success': False, 'message': 'No checklist template'}, status=404)
        
        checklist_template = execution.template.checklist_template
        
        # Get items
        items = []
        for item in checklist_template.items.all().order_by('no_urut'):
            items.append({
                'id': item.id,
                'item_pemeriksaan': item.item_pemeriksaan,
                'standar_normal': item.standar_normal,
                'satuan': item.unit,
                'unit': item.unit,
                'nilai_min': float(item.nilai_min) if item.nilai_min else None,
                'nilai_max': float(item.nilai_max) if item.nilai_max else None,
                'tindakan_remark': item.tindakan_remark,
            })
        
        # Get existing result (if any)
        existing_result = None
        try:
            checklist_result = ChecklistResult.objects.get(execution=execution)
            
            # Handle diisi_oleh - could be NULL for submissions via WhatsApp share link
            if checklist_result.diisi_oleh:
                diisi_oleh_name = checklist_result.diisi_oleh.get_full_name() or checklist_result.diisi_oleh.username
                diisi_oleh_username = checklist_result.diisi_oleh.username
            else:
                # For WhatsApp submissions, use accessed_by_name
                diisi_oleh_name = checklist_result.accessed_by_name or 'Unknown'
                diisi_oleh_username = ''
            
            existing_result = {
                'id': checklist_result.id,
                'hasil_pengukuran': checklist_result.hasil_pengukuran or {},
                'catatan_umum': checklist_result.catatan_umum or '',
                'status_overall': checklist_result.status_overall,
                'diisi_oleh': diisi_oleh_name,
                'diisi_oleh_username': diisi_oleh_username,
                'tanggal_pengisian': checklist_result.tanggal_pengisian.strftime('%d/%m/%Y %H:%M') if checklist_result.tanggal_pengisian else '',
            }
        except ChecklistResult.DoesNotExist:
            pass
        
        return JsonResponse({
            'success': True,
            'items': items,
            'existing_result': existing_result,
            'checklist_template_name': checklist_template.nama,
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


# ------------------------------------------------------------------------------
# SHAREABLE LINK: generate token + link to share via WA or other channel
# ------------------------------------------------------------------------------
@login_required(login_url='core:login')
@require_http_methods(["POST"])
def generate_share_link(request, execution_id):
    """Generate a time-limited token and shareable link for an execution's checklist.
    Returns JSON with `link` and `message` (text ready to paste to WA).
    """
    try:
        execution = get_object_or_404(PreventiveJobExecution, pk=execution_id)

        # Basic permission: only PIC, assigned, or staff can create share link
        user = request.user
        pic = execution.template.pic if execution.template else None
        if not (user.is_staff or user.is_superuser or user == execution.assigned_to or user == pic):
            return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)

        signer = signing.TimestampSigner(salt=SHARE_SIGN_SALT)
        payload = {'execution_id': execution.id}
        token = signing.dumps(payload, salt=SHARE_SIGN_SALT)

        scheme = request.scheme
        host = request.get_host()
        link = f"{scheme}://{host}{request.path.rsplit('/', 3)[0]}/checklist-fill/{token}/"

        message = f"Silakan isi checklist untuk pekerjaan: {execution.template.nama_pekerjaan} \nLink: {link}"

        return JsonResponse({'success': True, 'link': link, 'message': message, 'token': token})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@require_http_methods(["GET"])
def checklist_fill_view(request, token):
    """Render a public checklist fill page based on a signed token.
    Token contains execution_id and is time-limited.
    """
    try:
        data = signing.loads(token, salt=SHARE_SIGN_SALT, max_age=SHARE_TOKEN_MAX_AGE)
        execution_id = data.get('execution_id')
    except signing.SignatureExpired:
        return render(request, 'preventive_jobs/checklist_fill.html', {'error': 'Link telah kadaluarsa.'})
    except Exception:
        return render(request, 'preventive_jobs/checklist_fill.html', {'error': 'Token tidak valid.'})

    execution = get_object_or_404(PreventiveJobExecution, pk=execution_id)

    # Load checklist items and any existing result server-side to avoid auth issues
    checklist_template = execution.template.checklist_template if execution.template else None
    items = []
    existing_result = None
    if checklist_template:
        for item in checklist_template.items.all().order_by('no_urut'):
            items.append({
                'id': item.id,
                'item_pemeriksaan': item.item_pemeriksaan,
                'standar_normal': item.standar_normal,
                'satuan': item.unit,
            })
        try:
            from .models import ChecklistResult
            cr = ChecklistResult.objects.filter(execution=execution).first()
            if cr:
                existing_result = {
                    'id': cr.id,
                    'hasil_pengukuran': cr.hasil_pengukuran or {},
                    'catatan_umum': cr.catatan_umum or '',
                    'status_overall': cr.status_overall,
                }
        except Exception:
            existing_result = None

    context = {
        'execution': execution,
        'items': items,
        'existing_result': existing_result,
        'token': token,
    }
    return render(request, 'preventive_jobs/checklist_fill.html', context)


@csrf_exempt
@require_http_methods(["POST"])
def save_checklist_via_token(request):
    """Save checklist result submitted via token-based public form.
    Expected JSON body: { token: str, hasil_pengukuran: {...}, status_item: {...}, catatan_umum: str, nama: str (optional), nomor: str (optional) }
    """
    try:
        body = json.loads(request.body)
    except Exception:
        return JsonResponse({'success': False, 'message': 'Invalid JSON'}, status=400)

    token = body.get('token')
    if not token:
        return JsonResponse({'success': False, 'message': 'Missing token'}, status=400)

    try:
        data = signing.loads(token, salt=SHARE_SIGN_SALT, max_age=SHARE_TOKEN_MAX_AGE)
        execution_id = data.get('execution_id')
    except signing.SignatureExpired:
        return JsonResponse({'success': False, 'message': 'Token expired'}, status=400)
    except Exception:
        return JsonResponse({'success': False, 'message': 'Invalid token'}, status=400)

    execution = get_object_or_404(PreventiveJobExecution, pk=execution_id)

    # Parse submissions
    hasil_pengukuran = body.get('hasil_pengukuran', {})
    status_item = body.get('status_item', {})
    catatan = body.get('catatan_umum', '')
    nama = body.get('nama', '')
    nomor = body.get('nomor', '')

    # Convert simple values to expected format (same logic as original save)
    converted_hasil = {}
    for item_id, nilai in hasil_pengukuran.items():
        if isinstance(nilai, dict):
            converted_hasil[item_id] = nilai
        else:
            converted_hasil[item_id] = {
                'nilai': nilai,
                'status': status_item.get(item_id, 'OK')
            }

    if not converted_hasil:
        return JsonResponse({'success': False, 'message': 'No hasil_pengukuran provided'}, status=400)

    # Save ChecklistResult (diisi_oleh left as NULL)
    from .models import ChecklistResult
    checklist_result, created = ChecklistResult.objects.get_or_create(execution=execution)
    checklist_result.checklist_template = execution.template.checklist_template
    checklist_result.hasil_pengukuran = converted_hasil
    checklist_result.status_item = status_item
    if catatan:
        checklist_result.catatan = catatan
        checklist_result.catatan_umum = catatan
    # record submitter info into catatan_umum if provided
    submitter_info = []
    if nama:
        submitter_info.append(f"Nama: {nama}")
    if nomor:
        submitter_info.append(f"Nomor: {nomor}")
    if submitter_info:
        append = " | ".join(submitter_info)
        checklist_result.catatan_umum = (checklist_result.catatan_umum or '') + "\n" + append

    checklist_result.diisi_oleh = None
    if created:
        checklist_result.tanggal_pengisian = timezone.now()

    # Determine overall status
    status_overall = 'OK'
    for _, st in status_item.items():
        if st == 'NG':
            status_overall = 'NG'
            break
    checklist_result.status_overall = status_overall
    checklist_result.save()

    # Optionally update execution status to Done
    try:
        execution.status = 'Done'
        execution.save(update_fields=['status', 'updated_at'])
    except Exception:
        pass

    return JsonResponse({'success': True, 'message': 'Checklist saved via token', 'checklist_id': checklist_result.id})


@login_required(login_url='core:login')
@require_http_methods(["GET"])
def checklist_modal_api(request, execution_id):
    """
    AJAX endpoint untuk load checklist items untuk modal
    Format disesuaikan dengan checklist-modal.js yang support numeric dan text items
    Accessible by: assigned user, PIC, or superior of PIC
    """
    from .models import PreventiveJobExecution, ChecklistResult
    
    try:
        execution = get_object_or_404(PreventiveJobExecution, pk=execution_id)
        
        # Get PIC and check hierarchy
        pic = execution.template.pic
        current_user = request.user
        
        # Check permission - allow if:
        # 1. User is the assigned person
        # 2. User is the PIC
        # 3. User is in PIC's chain of command (atasan)
        is_allowed = False
        
        if current_user == execution.assigned_to:
            is_allowed = True
        elif current_user == pic:
            is_allowed = True
        elif pic and hasattr(pic, 'atasan') and pic.atasan:
            # Check if current user is in the superior chain
            superior = pic.atasan
            while superior:
                if current_user == superior:
                    is_allowed = True
                    break
                superior = superior.atasan if hasattr(superior, 'atasan') else None
        
        if not is_allowed:
            return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)
        
        # Get checklist template
        if not execution.template.checklist_template:
            return JsonResponse({'status': 'error', 'message': 'No checklist template'}, status=404)
        
        checklist_template = execution.template.checklist_template
        
        # Get items dengan support untuk numeric dan text
        items = []
        for item in checklist_template.items.all().order_by('no_urut'):
            item_data = {
                'id': item.id,
                'item_pemeriksaan': item.item_pemeriksaan,
                'standar_normal': item.standar_normal,
                'unit': item.unit or '',
                'nilai_min': float(item.nilai_min) if item.nilai_min else None,
                'nilai_max': float(item.nilai_max) if item.nilai_max else None,
                'item_type': item.item_type or 'numeric',
                'text_options': item.text_options or '',
                'tindakan_remark': item.tindakan_remark or '',
            }
            items.append(item_data)
        
        # Get existing result (if any)
        checklist_result = None
        try:
            result = ChecklistResult.objects.get(execution=execution)
            
            # Handle diisi_oleh - could be NULL for submissions via WhatsApp share link
            if result.diisi_oleh:
                diisi_oleh_name = result.diisi_oleh.get_full_name() or result.diisi_oleh.username
                diisi_oleh_username = result.diisi_oleh.username
            else:
                # For WhatsApp submissions, use accessed_by_name
                diisi_oleh_name = result.accessed_by_name or 'Unknown'
                diisi_oleh_username = ''
            
            checklist_result = {
                'id': result.id,
                'hasil_pengukuran': result.hasil_pengukuran or {},
                'status_item': result.status_item or {},
                'catatan': result.catatan or '',
                'status_overall': result.status_overall,
                'diisi_oleh': diisi_oleh_name,
                'diisi_oleh_username': diisi_oleh_username,
                'tanggal_pengisian': result.tanggal_pengisian.strftime('%d/%m/%Y %H:%M') if result.tanggal_pengisian else '',
            }
        except ChecklistResult.DoesNotExist:
            pass
        
        return JsonResponse({
            'status': 'success',
            'items': items,
            'checklist_result': checklist_result,
            'checklist_template_name': checklist_template.nama,
        })
    
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)


@login_required(login_url='core:login')
@require_http_methods(["POST"])
def save_checklist_result_view(request, execution_id):
    """
    AJAX endpoint untuk save hasil checklist
    Accessible by: assigned user, PIC, or superior of PIC
    """
    from .models import PreventiveJobExecution, ChecklistResult
    import json
    
    try:
        execution = get_object_or_404(PreventiveJobExecution, pk=execution_id)
        
        # Get PIC and check hierarchy
        pic = execution.template.pic
        current_user = request.user
        
        # Check permission - allow if:
        # 1. User is the assigned person
        # 2. User is the PIC
        # 3. User is in PIC's chain of command (atasan)
        is_allowed = False
        
        if current_user == execution.assigned_to:
            is_allowed = True
        elif current_user == pic:
            is_allowed = True
        elif pic and hasattr(pic, 'atasan') and pic.atasan:
            # Check if current user is in the superior chain
            superior = pic.atasan
            while superior:
                if current_user == superior:
                    is_allowed = True
                    break
                superior = superior.atasan if hasattr(superior, 'atasan') else None
        
        if not is_allowed:
            return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)
        
        # Get checklist template
        if not execution.template.checklist_template:
            return JsonResponse({'success': False, 'message': 'No checklist template'}, status=404)
        
        # Parse request body
        try:
            body = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON'}, status=400)
        
        # Handle both old and new format
        hasil_pengukuran = body.get('hasil_pengukuran', {})
        status_item = body.get('status_item', {})
        catatan = body.get('catatan', body.get('catatan_umum', ''))
        
        # If hasil_pengukuran is in new format (simple values), convert to old format
        converted_hasil = {}
        for item_id, nilai in hasil_pengukuran.items():
            if isinstance(nilai, dict):
                # Already in correct format
                converted_hasil[item_id] = nilai
            else:
                # Convert simple value to dict format
                converted_hasil[item_id] = {
                    'nilai': nilai,
                    'status': status_item.get(item_id, 'OK')
                }
        
        if not converted_hasil:
            return JsonResponse({'success': False, 'message': 'No hasil pengukuran provided'}, status=400)
        
        # Get or create ChecklistResult
        checklist_result, created = ChecklistResult.objects.get_or_create(
            execution=execution
        )
        
        # Update fields
        checklist_result.checklist_template = execution.template.checklist_template
        checklist_result.hasil_pengukuran = converted_hasil
        checklist_result.status_item = status_item
        # Store catatan in both fields for backwards compatibility
        if catatan:
            checklist_result.catatan = catatan
            checklist_result.catatan_umum = catatan
        checklist_result.diisi_oleh = request.user
        # Only update tanggal_pengisian if first time creation
        if created:
            checklist_result.tanggal_pengisian = timezone.now()
        
        # Determine status_overall based on status_item
        status_overall = 'OK'
        if status_item:
            for item_id, status in status_item.items():
                if status == 'NG':
                    status_overall = 'NG'
                    break
        
        checklist_result.status_overall = status_overall
        checklist_result.save()
        
        # AUTO SET EXECUTION STATUS TO DONE saat checklist diisi
        execution.status = 'Done'
        execution.save(update_fields=['status', 'updated_at'])
        
        return JsonResponse({
            'success': True,
            'message': 'Checklist result saved successfully',
            'checklist_id': checklist_result.id,
        })
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required(login_url='core:login')
@require_http_methods(['POST'])
def delete_checklist_result_view(request, execution_id):
    """
    AJAX endpoint untuk delete/reset hasil checklist
    Sehingga status execution bisa dirubah kembali
    """
    from .models import PreventiveJobExecution, ChecklistResult
    
    try:
        execution = get_object_or_404(PreventiveJobExecution, pk=execution_id)
        
        # Check permission
        if not (request.user == execution.assigned_to or 
                request.user in execution.template.pic.get_all_subordinates() or
                request.user == execution.template.pic):
            return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)
        
        # Delete checklist result if exists
        ChecklistResult.objects.filter(execution=execution).delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Checklist result deleted successfully'
        })
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


# ==============================================================================
# CHECKLIST MANAGEMENT VIEWS (WEB UI)
# ==============================================================================

@login_required(login_url='core:login')
def checklist_management_view(request):
    """
    Halaman manajemen checklist templates dan items
    """
    from .models import ChecklistTemplate
    
    # Get filter parameters
    search_query = request.GET.get('q', '')
    kategori = request.GET.get('kategori', '')
    status_filter = request.GET.get('status', '')
    
    # Get templates with filters
    templates = ChecklistTemplate.objects.all()
    
    if search_query:
        templates = templates.filter(
            Q(nama__icontains=search_query) | 
            Q(nomor__icontains=search_query) |
            Q(deskripsi__icontains=search_query)
        )
    
    if kategori:
        templates = templates.filter(kategori=kategori)
    
    if status_filter == 'active':
        templates = templates.filter(is_active=True)
    elif status_filter == 'inactive':
        templates = templates.filter(is_active=False)
    
    templates = templates.order_by('-created_at')
    
    context = {
        'templates': templates,
        'search_query': search_query,
        'kategori': kategori,
        'status': status_filter,
    }
    
    return render(request, 'preventive_jobs/checklist_management.html', context)


@login_required(login_url='core:login')
@require_http_methods(['POST'])
def checklist_template_create_api(request):
    """
    API endpoint untuk create checklist template
    """
    from .models import ChecklistTemplate
    
    try:
        nama = request.POST.get('nama')
        kategori = request.POST.get('kategori')
        deskripsi = request.POST.get('deskripsi', '')
        is_active = request.POST.get('is_active') == 'on'
        
        if not nama or not kategori:
            return JsonResponse({'success': False, 'message': 'Nama dan kategori wajib diisi'}, status=400)
        
        template = ChecklistTemplate.objects.create(
            nama=nama,
            kategori=kategori,
            deskripsi=deskripsi,
            is_active=is_active,
            created_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Template berhasil dibuat',
            'template_id': template.id
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required(login_url='core:login')
@require_http_methods(['POST', 'GET'])
def checklist_template_api(request, pk):
    """
    API endpoint untuk get/delete checklist template
    """
    from .models import ChecklistTemplate
    
    try:
        template = get_object_or_404(ChecklistTemplate, pk=pk)
        
        if request.method == 'GET':
            return JsonResponse({
                'success': True,
                'template': {
                    'id': template.id,
                    'nama': template.nama,
                    'kategori': template.kategori,
                    'deskripsi': template.deskripsi,
                    'is_active': template.is_active,
                }
            })
        
        elif request.method == 'POST' and 'delete' in request.path:
            # Delete endpoint
            template.delete()
            return JsonResponse({'success': True, 'message': 'Template dihapus'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required(login_url='core:login')
@require_http_methods(['POST'])
def checklist_template_update_api(request):
    """
    API endpoint untuk update checklist template
    """
    from .models import ChecklistTemplate
    
    try:
        template_id = request.POST.get('template_id')
        nama = request.POST.get('nama')
        kategori = request.POST.get('kategori')
        deskripsi = request.POST.get('deskripsi', '')
        is_active = request.POST.get('is_active') == 'on'
        
        if not template_id:
            return JsonResponse({'success': False, 'message': 'Template ID wajib diisi'}, status=400)
        
        template = get_object_or_404(ChecklistTemplate, pk=template_id)
        template.nama = nama
        template.kategori = kategori
        template.deskripsi = deskripsi
        template.is_active = is_active
        template.save()
        
        return JsonResponse({'success': True, 'message': 'Template berhasil diupdate'})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required(login_url='core:login')
@require_http_methods(['GET', 'POST'])
def checklist_items_api(request, pk):
    """
    API endpoint untuk get items dari template tertentu
    """
    from .models import ChecklistTemplate
    
    try:
        template = get_object_or_404(ChecklistTemplate, pk=pk)
        
        if request.method == 'GET':
            items = []
            for item in template.items.all().order_by('no_urut'):
                items.append({
                    'id': item.id,
                    'no_urut': item.no_urut,
                    'item_type': item.item_type or 'numeric',  # ADD THIS
                    'item_pemeriksaan': item.item_pemeriksaan,
                    'standar_normal': item.standar_normal,
                    'unit': item.unit or '',
                    'nilai_min': float(item.nilai_min) if item.nilai_min else None,
                    'nilai_max': float(item.nilai_max) if item.nilai_max else None,
                    'text_options': item.text_options or '',  # ADD THIS
                    'tindakan_remark': item.tindakan_remark or '',
                })
            
            return JsonResponse({
                'success': True,
                'template_name': template.nama,
                'items': items
            })
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required(login_url='core:login')
@require_http_methods(['POST'])
def checklist_items_save_api(request):
    """
    API endpoint untuk save/update checklist items
    """
    from .models import ChecklistTemplate, ChecklistItem
    import logging
    
    logger = logging.getLogger(__name__)
    
    try:
        body = json.loads(request.body)
        template_id = body.get('template_id')
        items_data = body.get('items', [])
        
        # Log what we receive (using proper logging instead of print)
        logger.info(f"[SAVE ITEMS API] Received {len(items_data)} items for template {template_id}")
        for idx, item in enumerate(items_data):
            logger.debug(f"  Item {idx+1}: {item.get('item_pemeriksaan')} (type: {item.get('item_type')})")
        
        if not template_id:
            return JsonResponse({'success': False, 'message': 'Template ID wajib diisi'}, status=400)
        
        template = get_object_or_404(ChecklistTemplate, pk=template_id)
        
        # Delete existing items
        template.items.all().delete()
        
        # Create new items
        for item_data in items_data:
            item_type = item_data.get('item_type', 'numeric')
            logger.debug(f"[CREATE] Saving item: {item_data.get('item_pemeriksaan')} with type: {item_type}")
            
            ChecklistItem.objects.create(
                checklist_template=template,
                no_urut=item_data.get('no_urut'),
                item_type=item_type,
                item_pemeriksaan=item_data.get('item_pemeriksaan'),
                standar_normal=item_data.get('standar_normal'),
                unit=item_data.get('unit') or '',
                nilai_min=item_data.get('nilai_min'),
                nilai_max=item_data.get('nilai_max'),
                text_options=item_data.get('text_options', ''),
                tindakan_remark=item_data.get('tindakan_remark', ''),
            )
        
        logger.info(f"Successfully saved {len(items_data)} items for template {template_id}")
        
        return JsonResponse({
            'success': True,
            'message': f'{len(items_data)} items berhasil disimpan'
        })
    
    except json.JSONDecodeError as e:
        logger.error(f"JSON decode error: {str(e)}")
        return JsonResponse({'success': False, 'message': 'Invalid JSON format'}, status=400)
    except Exception as e:
        logger.error(f"Error saving checklist items: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)


@login_required(login_url='core:login')
@require_http_methods(['GET'])
def checklist_template_download_excel(request, pk):
    """
    Download template Excel untuk import items
    """
    from .models import ChecklistTemplate, ChecklistItem
    
    try:
        template = get_object_or_404(ChecklistTemplate, pk=pk)
        
        # Try to use openpyxl, fallback to xlsxwriter
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Items"
            
            # Set column widths
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 20
            
            # Header dengan Item Type dan Pilihan Text
            headers = ['No', 'Item Type', 'Item Pemeriksaan', 'Standar/Normal', 'Unit', 'Min', 'Max', 'Pilihan Text', 'Tindakan/Remark']
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data rows with existing items
            items = template.items.all().order_by('no_urut')
            for idx, item in enumerate(items, 1):
                ws.cell(row=idx+1, column=1, value=idx)
                ws.cell(row=idx+1, column=2, value=item.item_type or 'numeric')
                ws.cell(row=idx+1, column=3, value=item.item_pemeriksaan)
                ws.cell(row=idx+1, column=4, value=item.standar_normal)
                ws.cell(row=idx+1, column=5, value=item.unit or '')
                ws.cell(row=idx+1, column=6, value=item.nilai_min or '')
                ws.cell(row=idx+1, column=7, value=item.nilai_max or '')
                ws.cell(row=idx+1, column=8, value=item.text_options or '')
                ws.cell(row=idx+1, column=9, value=item.tindakan_remark or '')
            
            # Add empty rows for new data
            for idx in range(len(items) + 1, len(items) + 11):
                ws.cell(row=idx+1, column=1, value=idx)
            
            # Save to response
            from io import BytesIO
            from django.http import FileResponse
            
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            response = FileResponse(excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="Checklist-{template.nomor}-Items.xlsx"'
            return response
            
        except ImportError:
            # Fallback to CSV if openpyxl not available
            import csv
            from django.http import HttpResponse
            
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="Checklist-{template.nomor}-Items.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['No', 'Item Pemeriksaan', 'Standar/Normal', 'Unit', 'Min', 'Max', 'Tindakan/Remark'])
            
            items = template.items.all().order_by('no_urut')
            for idx, item in enumerate(items, 1):
                writer.writerow([
                    idx,
                    item.item_pemeriksaan,
                    item.standar_normal,
                    item.unit or '',
                    item.nilai_min or '',
                    item.nilai_max or '',
                    item.tindakan_remark or ''
                ])
            
            return response
            
    except Exception as e:
        from django.http import HttpResponse
        return HttpResponse(f'Error: {str(e)}', status=500)


@login_required(login_url='core:login')
@require_http_methods(['POST'])
def checklist_items_import_excel(request):
    """
    Import items dari Excel file
    """
    from .models import ChecklistTemplate, ChecklistItem
    
    try:
        template_id = request.POST.get('template_id')
        excel_file = request.FILES.get('file')
        
        if not template_id or not excel_file:
            return JsonResponse({'success': False, 'message': 'Template ID dan file wajib diisi'}, status=400)
        
        template = get_object_or_404(ChecklistTemplate, pk=template_id)
        
        # Read Excel file
        try:
            from openpyxl import load_workbook
            from io import BytesIO
            
            # Load workbook
            wb = load_workbook(BytesIO(excel_file.read()))
            ws = wb.active
            
            # Skip header row
            items_to_create = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
                if not row or not row[2]:  # Skip if no Item Pemeriksaan (now column 3, was column 2)
                    continue
                
                # Determine item_type from column B (column index 1)
                item_type = str(row[1] or 'numeric').lower().strip() if row[1] else 'numeric'
                if item_type not in ['numeric', 'text']:
                    item_type = 'numeric'  # Default if invalid
                
                items_to_create.append({
                    'no_urut': row_idx,
                    'item_type': item_type,
                    'item_pemeriksaan': str(row[2] or ''),
                    'standar_normal': str(row[3] or ''),
                    'unit': str(row[4] or '') if row[4] else None,
                    'nilai_min': float(row[5]) if row[5] and str(row[5]).strip() else None,
                    'nilai_max': float(row[6]) if row[6] and str(row[6]).strip() else None,
                    'text_options': str(row[7] or '') if row[7] else '',
                    'tindakan_remark': str(row[8] or '') if row[8] else None,
                })
            
            # Delete existing items
            template.items.all().delete()
            
            # Create new items
            for item_data in items_to_create:
                ChecklistItem.objects.create(
                    checklist_template=template,
                    **item_data
                )
            
            return JsonResponse({
                'success': True,
                'message': f'{len(items_to_create)} items berhasil diimport dari Excel'
            })
        
        except ImportError:
            # Fallback to CSV reading
            import csv
            from io import TextIOWrapper
            
            items_to_create = []
            csv_file = TextIOWrapper(excel_file, encoding='utf-8')
            reader = csv.reader(csv_file)
            next(reader)  # Skip header
            
            for row_idx, row in enumerate(reader, start=1):
                if not row or len(row) < 3 or not row[2]:  # Check column 3 (Item Pemeriksaan)
                    continue
                
                # Determine item_type from column B (index 1)
                item_type = str(row[1] or 'numeric').lower().strip() if len(row) > 1 and row[1] else 'numeric'
                if item_type not in ['numeric', 'text']:
                    item_type = 'numeric'
                
                items_to_create.append({
                    'no_urut': row_idx,
                    'item_type': item_type,
                    'item_pemeriksaan': str(row[2] or ''),
                    'standar_normal': str(row[3] or '') if len(row) > 3 else '',
                    'unit': str(row[4] or '') if len(row) > 4 and row[4] else None,
                    'nilai_min': float(row[5]) if len(row) > 5 and row[5] and str(row[5]).strip() else None,
                    'nilai_max': float(row[6]) if len(row) > 6 and row[6] and str(row[6]).strip() else None,
                    'text_options': str(row[7] or '') if len(row) > 7 and row[7] else '',
                    'tindakan_remark': str(row[8] or '') if len(row) > 8 and row[8] else None,
                })
            
            # Delete existing items
            template.items.all().delete()
            
            # Create new items
            for item_data in items_to_create:
                ChecklistItem.objects.create(
                    checklist_template=template,
                    **item_data
                )
            
            return JsonResponse({
                'success': True,
                'message': f'{len(items_to_create)} items berhasil diimport dari CSV'
            })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ==============================================================================
# TRIAL: JOB PER DAY UNTUK PREVENTIVE JOBS (HALAMAN BARU)
# ==============================================================================

@login_required(login_url='core:login')
def job_per_day_trial_view(request):
    """
    HALAMAN TRIAL: Menampilkan SEMUA jobs dalam format tabel (Daily + Project + Preventive)
    - Support filter: Date, Line, Mesin, Sub Mesin, Progress
    - Export ke berbagai form (Preventif, Evaluasi, dll)
    - TRIAL page (dipisah dari /job-per-day/ untuk testing)
    """
    import datetime
    from core.models import Job, JobDate
    
    user = request.user
    now = datetime.datetime.now()
    
    # === GET FILTER PARAMETERS ===
    selected_date_str = request.GET.get('date', now.strftime('%Y-%m-%d'))
    try:
        selected_date = datetime.datetime.strptime(selected_date_str, '%Y-%m-%d').date()
    except ValueError:
        selected_date = now.date()
    
    # === GET USER'S SUBORDINATES FOR PERMISSION CHECK ===
    subordinate_ids = user.get_all_subordinates()
    all_user_ids = [user.id] + subordinate_ids
    
    # === COLLECT ALL JOBS (DAILY + PROJECT + PREVENTIVE) ===
    all_jobs_unified = []
    
    # 1. DAILY & PROJECT JOBS
    jobs_for_date = JobDate.objects.filter(
        tanggal=selected_date
    ).select_related('job').prefetch_related(
        'job__personil_ditugaskan',
        'job__aset',
        'job__aset__parent',
        'job__aset__parent__parent'
    )
    
    # Filter berdasarkan permission
    for job_date in jobs_for_date:
        job = job_date.job
        can_view = (
            job.pic == user or
            job.assigned_to == user or
            (job.pic and job.pic.id in subordinate_ids) or
            (job.assigned_to is not None and job.assigned_to.id in subordinate_ids)
        )
        if can_view:
            # Wrap ke unified format
            all_jobs_unified.append({
                'type': 'daily' if job.tipe_job == 'Daily' else 'project',
                'tipe_display': job.get_tipe_job_display(),
                'id': job.id,
                'nama': job.nama_pekerjaan,
                'template_nama': '-',
                'line': job.aset.parent.parent.nama if (job.aset and job.aset.parent and job.aset.parent.parent) else '-',
                'mesin': job.aset.parent.nama if (job.aset and job.aset.parent) else '-',
                'submesin': job.aset.nama if job.aset else '-',
                'pic': job.pic.username if job.pic else '-',
                'assigned_to': job.assigned_to.username if job.assigned_to else '-',
                'personil': list(job.personil_ditugaskan.all()),
                'prioritas': job.get_prioritas_display(),
                'fokus': job.get_fokus_display(),
                'status': job_date.status,
                'catatan': '-',
                'progress': 100 if job_date.status == 'Selesai' else (50 if job_date.status == 'Proses' else 0),
            })
    
    # 2. PREVENTIVE JOBS
    executions_for_date = PreventiveJobExecution.objects.filter(
        scheduled_date=selected_date
    ).select_related(
        'template',
        'template__pic',
        'aset',
        'aset__parent',
        'aset__parent__parent',
        'assigned_to'
    ).prefetch_related('assigned_to_personil')
    
    # Filter berdasarkan permission
    for execution in executions_for_date:
        can_view = (
            execution.template.pic == user or
            execution.assigned_to == user or
            (execution.template.pic and execution.template.pic.id in subordinate_ids) or
            (execution.assigned_to and execution.assigned_to.id in subordinate_ids)
        )
        if can_view:
            # Wrap ke unified format
            all_jobs_unified.append({
                'type': 'preventive',
                'tipe_display': 'Preventive',
                'id': execution.id,
                'nama': execution.template.nama_pekerjaan,
                'template_nama': execution.template.nama_pekerjaan,
                'line': execution.aset.parent.parent.nama if (execution.aset and execution.aset.parent and execution.aset.parent.parent) else '-',
                'mesin': execution.aset.parent.nama if (execution.aset and execution.aset.parent) else '-',
                'submesin': execution.aset.nama if execution.aset else '-',
                'pic': execution.template.pic.username if execution.template.pic else '-',
                'assigned_to': execution.assigned_to.username if execution.assigned_to else '-',
                'personil': list(execution.assigned_to_personil.all()),
                'prioritas': execution.template.get_prioritas_display() if execution.template else '-',
                'fokus': execution.template.get_fokus_display() if execution.template else '-',
                'status': execution.status,
                'catatan': execution.catatan or '-',
                'progress': 100 if execution.status == 'Done' else 0,
            })
    
    # === COLLECT UNIQUE FILTER VALUES ===
    all_lines = set()
    all_mesins = set()
    all_submesins = set()
    all_progress = set()
    
    for job in all_jobs_unified:
        if job['line'] != '-':
            all_lines.add(job['line'])
        if job['mesin'] != '-':
            all_mesins.add(job['mesin'])
        if job['submesin'] != '-':
            all_submesins.add(job['submesin'])
        
        all_progress.add(f"{job['progress']}%")
    
    # Sort untuk dropdown
    all_lines = sorted(list(all_lines))
    all_mesins = sorted(list(all_mesins))
    all_submesins = sorted(list(all_submesins))
    all_progress = sorted(list(all_progress))
    
    context = {
        'selected_date': selected_date,
        'all_jobs': all_jobs_unified,
        'all_lines': all_lines,
        'all_mesins': all_mesins,
        'all_submesins': all_submesins,
        'all_progress': all_progress,
        'total_jobs': len(all_jobs_unified),
    }
    
    return render(request, 'preventive_jobs/job_per_day_trial.html', context)


# ==============================================================================
# EXPORT ENDPOINT UNTUK TRIAL PAGE (UNIFIED JOBS)
# ==============================================================================

@login_required(login_url='core:login')
@require_http_methods(['POST'])
def export_unified_jobs_to_gas(request, export_type=None):
    """
    Export unified jobs (Daily + Project + Preventive) ke Google Apps Script
    Supports: Preventif dan Evaluasi forms
    
    Args:
        export_type: Optional, dapat pass via URL parameter atau di request body
                    ("preventif" atau "evaluasi")
    """
    import json
    from core.export_handlers import prepare_job_data_for_export, send_to_google_apps_script
    from core.models import Job
    
    try:
        body = json.loads(request.body)
        job_ids_str = body.get('job_ids', [])  # Format: ["1_daily", "2_project", "3_preventive"]
        
        # Get export_type dari parameter atau request body (parameter takes precedence)
        if export_type is None:
            export_type = body.get('export_type', 'preventif')  # default: "preventif"
        
        if not job_ids_str or export_type not in ['preventif', 'evaluasi']:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid parameters'
            }, status=400)
        
        # Parse job IDs dari format "id_type"
        daily_job_ids = []
        project_job_ids = []
        preventive_job_ids = []
        
        for job_id_str in job_ids_str:
            try:
                job_id, job_type = job_id_str.rsplit('_', 1)
                job_id = int(job_id)
                
                if job_type == 'daily' or job_type == 'project':
                    daily_job_ids.append(job_id)  # Both are Job model
                elif job_type == 'preventive':
                    preventive_job_ids.append(job_id)
            except (ValueError, AttributeError):
                continue
        
        # === PREPARE DATA ===
        job_data_list = []
        
        # 1. Process Daily/Project jobs
        if daily_job_ids:
            jobs = Job.objects.filter(id__in=daily_job_ids).select_related(
                'aset', 'aset__parent', 'aset__parent__parent', 'pic', 'assigned_to'
            ).prefetch_related('personil_ditugaskan', 'tanggal_pelaksanaan')
            
            for job in jobs:
                personil_names = ", ".join([p.nama_lengkap for p in job.personil_ditugaskan.all()])
                mesin_name = job.aset.parent.nama if job.aset and job.aset.level == 2 and job.aset.parent else (job.aset.nama if job.aset else "")
                sub_mesin_name = job.aset.nama if job.aset and job.aset.level == 2 else ""
                
                if export_type == "preventif":
                    job_data_list.append([
                        personil_names,
                        mesin_name,
                        sub_mesin_name,
                        job.nama_pekerjaan
                    ])
                elif export_type == "evaluasi":
                    job_data_list.append([
                        personil_names,
                        mesin_name,
                        sub_mesin_name,
                        job.nama_pekerjaan,
                        ""  # Kolom kesimpulan/status kosong
                    ])
        
        # 2. Process Preventive jobs
        if preventive_job_ids:
            executions = PreventiveJobExecution.objects.filter(id__in=preventive_job_ids).select_related(
                'template', 'template__pic', 'aset', 'aset__parent', 'aset__parent__parent', 'assigned_to'
            ).prefetch_related('assigned_to_personil')
            
            for execution in executions:
                personil_names = ", ".join([p.nama_lengkap for p in execution.assigned_to_personil.all()])
                mesin_name = execution.aset.parent.nama if execution.aset and execution.aset.parent else ""
                sub_mesin_name = execution.aset.nama if execution.aset else ""
                
                if export_type == "preventif":
                    job_data_list.append([
                        personil_names,
                        mesin_name,
                        sub_mesin_name,
                        execution.template.nama_pekerjaan
                    ])
                elif export_type == "evaluasi":
                    job_data_list.append([
                        personil_names,
                        mesin_name,
                        sub_mesin_name,
                        execution.template.nama_pekerjaan,
                        ""  # Kolom kesimpulan/status kosong
                    ])
        
        if not job_data_list:
            return JsonResponse({
                'status': 'error',
                'message': 'No valid jobs to export'
            }, status=400)
        
        # === BUILD PAYLOAD ===
        mesin_set = set()
        sub_mesin_set = set()
        prioritas_set = set()
        
        for row in job_data_list:
            if row[1]:  # Mesin
                mesin_set.add(row[1])
            if row[2]:  # Sub Mesin
                sub_mesin_set.add(row[2])
        
        payload = {
            "exportType": export_type,
            "tanggal": str(timezone.now().date()),
            "allMesin": ", ".join(sorted(mesin_set)) if mesin_set else "",
            "allSubMesin": ", ".join(sorted(sub_mesin_set)) if sub_mesin_set else "",
            "allPrioritas": "",
            "teknisiBertugas": "PIC/Teknisi",
            "jobData": job_data_list
        }
        
        # === SEND TO GOOGLE APPS SCRIPT ===
        result = send_to_google_apps_script(payload, export_type)
        
        return JsonResponse(result)
    
    except json.JSONDecodeError:
        return JsonResponse({
            'status': 'error',
            'message': 'Invalid JSON'
        }, status=400)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)


# ==============================================================================
# STATE MACHINE & UNDO ENDPOINTS (TESTING PHASE)
# ==============================================================================

@login_required(login_url='core:login')
@require_http_methods(['POST'])
def execution_undo_status_change(request, execution_id):
    """
    AJAX endpoint untuk UNDO last status change.
    Berguna untuk 'salah klik' atau testing fase.
    Return: (success: bool, message: str, new_status: str)
    """
    from .models import PreventiveJobExecution
    
    try:
        execution = get_object_or_404(PreventiveJobExecution, pk=execution_id)
        
        # Permission check
        if not (request.user == execution.assigned_to or 
                request.user == execution.template.pic or
                request.user in execution.template.pic.get_all_subordinates()):
            return JsonResponse({
                'success': False,
                'message': 'Permission denied'
            }, status=403)
        
        # Perform undo
        success, message = execution.undo_last_status_change(changed_by=request.user)
        
        if success:
            return JsonResponse({
                'success': True,
                'message': message,
                'new_status': execution.status
            })
        else:
            return JsonResponse({
                'success': False,
                'message': message
            }, status=400)
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required(login_url='core:login')
@require_http_methods(['GET'])
def execution_status_history(request, execution_id):
    """
    AJAX endpoint untuk get full status change history.
    Berguna untuk audit trail dan transparency.
    """
    from .models import PreventiveJobExecution
    
    try:
        execution = get_object_or_404(PreventiveJobExecution, pk=execution_id)
        
        # Permission check
        if not (request.user == execution.assigned_to or 
                request.user == execution.template.pic or
                request.user in execution.template.pic.get_all_subordinates()):
            return JsonResponse({
                'success': False,
                'message': 'Permission denied'
            }, status=403)
        
        # Get status history
        history = []
        for log in execution.get_status_history().order_by('-changed_at'):
            history.append({
                'id': log.id,
                'from_status': log.from_status,
                'to_status': log.to_status,
                'reason': log.reason or '(No reason)',
                'changed_by': log.changed_by.get_full_name() or log.changed_by.username if log.changed_by else 'System',
                'changed_at': log.changed_at.strftime('%Y-%m-%d %H:%M:%S'),
                'timestamp_iso': log.changed_at.isoformat(),
            })
        
        return JsonResponse({
            'success': True,
            'execution_id': execution_id,
            'current_status': execution.status,
            'history': history,
            'total_changes': len(history)
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


# ==============================================================================
# ATTACHMENTS API ENDPOINT
# ==============================================================================

@login_required(login_url='core:login')
@require_http_methods(["GET"])
def get_execution_attachments(request, execution_id):
    """
    AJAX endpoint untuk get attachments dari execution
    Return JSON dengan daftar file attachments
    """
    from .models import PreventiveJobExecution
    
    
    def user_can_view_execution(user, execution):
        """Check if user can view execution and its attachments"""
        # Staff/Superuser can view everything
        if user.is_staff or user.is_superuser:
            print(f"[DEBUG] {user} can view - is staff/superuser")
            return True
        
        # Allow all authenticated users to view (permissive approach for now)
        # This can be made more restrictive later if needed
        print(f"[DEBUG] {user} allowed to view execution {execution.id} (permissive)")
        return True
    
    try:
        execution = get_object_or_404(PreventiveJobExecution, pk=execution_id)
        
        if not user_can_view_execution(request.user, execution):
            return JsonResponse({'success': False, 'message': 'Permission denied'}, status=403)
        
        # Get attachments
        attachments = []
        for att in execution.attachments.all():
            attachments.append({
                'id': att.id,
                'file': att.file.url,
                'file_name': att.file.name.split('/')[-1],
                'uploaded_at': att.uploaded_at.strftime('%d/%m/%Y %H:%M') if att.uploaded_at else 'N/A',
            })
        
        return JsonResponse({
            'success': True,
            'attachments': attachments,
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


# ==============================================================================
# WHATSAPP SHARE VIEWS
# ==============================================================================


@login_required(login_url='core:login')
@require_http_methods(["GET"])
def share_checklist_modal_view(request, execution_id):
    """
    Modal untuk share checklist via WhatsApp.
    Return: JSON dengan data recipient dan job info
    """
    from preventive_jobs.models import WhatsAppContact
    
    try:
        execution = get_object_or_404(PreventiveJobExecution, id=execution_id)
        
        # Check apakah job punya checklist template
        if not execution.template.checklist_template:
            return JsonResponse({
                'success': False,
                'message': 'Job preventive ini tidak memiliki checklist template. Silakan konfigurasi checklist terlebih dahulu di admin panel.'
            }, status=400)
        
        # Check authorization
        user = request.user
        pic = execution.template.pic
        
        can_share = (user.id == pic.id)
        if not can_share:
            current = pic
            max_iter = 10
            while current and max_iter > 0:
                if hasattr(current, 'atasan') and current.atasan and current.atasan.id == user.id:
                    can_share = True
                    break
                current = current.atasan if hasattr(current, 'atasan') else None
                max_iter -= 1
        
        if not can_share:
            return JsonResponse({
                'success': False,
                'message': 'Anda tidak memiliki izin untuk share checklist ini'
            }, status=403)
        
        # Get available recipients
        recipients = []
        
        # Add PIC as default recipient
        if pic.nomor_telepon:
            recipients.append({
                'id': f'user_{pic.id}',
                'nama': pic.get_full_name() or pic.username,
                'nomor_wa': pic.nomor_telepon,
                'tipe': 'Personal (PIC)',
                'is_default': True,
            })
        
        # Add other WhatsApp contacts
        contacts = WhatsAppContact.objects.filter(is_active=True).order_by('tipe_kontak', 'nama')
        for contact in contacts:
            recipients.append({
                'id': f'contact_{contact.id}',
                'nama': contact.nama,
                'nomor_wa': contact.nomor_wa,
                'tipe': contact.get_tipe_kontak_display(),
                'is_default': False,
            })
        
        return JsonResponse({
            'success': True,
            'execution_id': execution_id,
            'job_name': execution.template.nama_pekerjaan,
            'aset_name': execution.aset.nama if execution.aset else 'N/A',
            'scheduled_date': execution.scheduled_date.strftime('%d-%m-%Y'),
            'recipients': recipients,
        })
    
    except PreventiveJobExecution.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Execution tidak ditemukan'
        }, status=404)
    
    except Exception as e:
        logger.error(f"Error share_checklist_modal_view: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}'
        }, status=500)


@login_required
def generate_share_link(request, execution_id):
    """
    Generate share link untuk checklist dan kirim via WhatsApp (via Fontte API).
    
    POST Data:
        - contact_ids: JSON array dari contact IDs
        - custom_message: (optional) pesan tambahan
    
    Returns:
        JSON response dengan status pengiriman
    """
    from preventive_jobs.whatsapp_utils import (
        generate_checklist_share_token,
        build_share_url,
        build_whatsapp_message,
        shorten_url,
        FontteAPI
    )
    from preventive_jobs.models import ChecklistShareLog, WhatsAppContact
    
    try:
        execution = get_object_or_404(PreventiveJobExecution, id=execution_id)
        
        # Check authorization
        user = request.user
        pic = execution.template.pic
        
        # PENTING: Validate checklist_template ada dulu!
        if not execution.template or not execution.template.checklist_template:
            return JsonResponse({
                'success': False,
                'message': '❌ SETUP ERROR: Template job ini tidak punya Checklist Template yang ter-link. Hubungi admin untuk setup checklist terlebih dahulu.'
            }, status=400)
        
        # User harus PIC atau atasan dari PIC
        can_share = (user.id == pic.id)
        if not can_share:
            # Check if user is superior
            current = pic
            max_iter = 10
            while current and max_iter > 0:
                if hasattr(current, 'atasan') and current.atasan and current.atasan.id == user.id:
                    can_share = True
                    break
                current = current.atasan if hasattr(current, 'atasan') else None
                max_iter -= 1
        
        if not can_share:
            return JsonResponse({
                'success': False,
                'message': 'Anda tidak memiliki izin untuk share checklist ini'
            }, status=403)
        
        # Get or create checklist result
        # PENTING: Deteksi apakah ini pengisian pertama atau edit
        checklist_result = ChecklistResult.objects.filter(execution=execution).first()
        is_first_submission = False
        warning_message = None
        
        if not checklist_result:
            # Pengisian PERTAMA - jangan buat di sini, buat saat user buka form
            # Ini agar status tidak berubah dari "Belum Diisi" ke "Partial"
            is_first_submission = True
            # Placeholder untuk store di token
            checklist_result = None
        else:
            # Edit data yang sudah ada (re-check)
            if checklist_result.status_overall in ['OK', 'NG']:
                warning_message = f"⚠️ EDIT MODE: Ini bukan pengisian pertama. Status saat ini adalah '{checklist_result.status_overall}'. Data checklist akan di-update."
            is_first_submission = False
        
        # Parse recipients from POST/JSON body
        import json
        
        # Try to parse JSON body first (for fetch POST)
        try:
            body = json.loads(request.body)
            contact_ids = body.get('contact_ids', [])
            custom_message = body.get('custom_message', '').strip()
        except:
            # Fallback to POST data
            recipients_json = request.POST.get('recipients', '[]')
            contact_ids = json.loads(recipients_json)
            custom_message = request.POST.get('custom_message', '').strip()
        
        # Build recipients list from contact IDs
        recipients = []
        
        for contact_id in contact_ids:
            try:
                if contact_id.startswith('user_'):
                    # PIC user
                    user_id = int(contact_id.split('_')[1])
                    contact_user = CustomUser.objects.get(id=user_id)
                    if contact_user.nomor_telepon:
                        recipients.append({
                            'nama': contact_user.get_full_name() or contact_user.username,
                            'nomor_wa': contact_user.nomor_telepon,
                            'type': 'user'
                        })
                elif contact_id.startswith('contact_'):
                    # WhatsApp contact
                    contact_pk = int(contact_id.split('_')[1])
                    contact = WhatsAppContact.objects.get(id=contact_pk)
                    recipients.append({
                        'nama': contact.nama,
                        'nomor_wa': contact.nomor_wa,
                        'type': 'contact'
                    })
            except Exception as e:
                logger.warning(f"Error parsing contact {contact_id}: {str(e)}")
                continue
        
        if not recipients:
            return JsonResponse({
                'success': False,
                'message': 'Tidak ada penerima yang valid'
            }, status=400)
        
        send_via_api = request.POST.get('send_via_api', 'true').lower() == 'true'
        
        # Check if API is enabled
        api_enabled = getattr(settings, 'FONTTE_API_ENABLED', False)
        if not api_enabled:
            send_via_api = False  # Force to False if API disabled
        
        # PENTING: JANGAN create ChecklistResult saat share!
        # ChecklistResult akan dibuat saat penerima BUKA FORM
        # Ini agar status tetap "Belum Diisi" sampai form dibuka
        if not checklist_result:
            # Ambil checklist_template dari execution.template
            template_obj = execution.template
            checklist_template_to_use = template_obj.checklist_template if template_obj else None
            
            if not checklist_template_to_use:
                return JsonResponse({
                    'success': False,
                    'message': '❌ ERROR: Checklist template tidak ter-link ke job template "' + (template_obj.nama_pekerjaan if template_obj else 'Unknown') + '". Hubungi administrator untuk setup checklist template.'
                }, status=400)
            
            # Untuk token generation, kita perlu checklist_result ID
            # Buat dengan status 'Pending' - ini hanya marker internal
            # Saat penerima buka form, status berubah Pending -> Partial
            checklist_result = ChecklistResult.objects.create(
                execution=execution,
                checklist_template=checklist_template_to_use,
                status_overall='Pending'  # Status khusus: share link dibuat tapi belum dibuka
            )
        
        # Generate share links for each recipient
        results = []
        fontte = FontteAPI() if api_enabled else None
        
        for recipient in recipients:
            try:
                nama = recipient.get('nama', 'Unknown').strip()[:255]  # Truncate to max_length
                nomor_wa = recipient.get('nomor_wa', '').strip()
                
                if not nomor_wa:
                    results.append({
                        'nama': nama,
                        'nomor_wa': nomor_wa,
                        'success': False,
                        'message': 'Nomor WA tidak valid'
                    })
                    continue
                
                # Generate token
                token = generate_checklist_share_token(execution.id, checklist_result.id)
                share_url = build_share_url(token)
                
                # Shorten URL untuk WhatsApp message (TinyURL)
                short_url = shorten_url(share_url)
                
                # Build message dengan short URL
                message = build_whatsapp_message(
                    execution,
                    checklist_result,
                    short_url,  # Use shortened URL
                    custom_message
                )
                
                # Create share log
                share_log = ChecklistShareLog.objects.create(
                    execution=execution,
                    checklist_result=checklist_result,
                    penerima_nama=nama,  # Already truncated
                    penerima_wa=nomor_wa[:20],  # Truncate to max_length
                    share_link=share_url[:200],  # URLField default max_length
                    share_token=token[:255],  # Truncate to max_length
                    pesan_wa=message,
                    dikirim_oleh=user,
                    status_pengiriman='pending'
                )
                
                # Send via API if requested and enabled
                if send_via_api and fontte:
                    api_result = fontte.send_message(nomor_wa, message)
                    
                    if api_result['success']:
                        share_log.status_pengiriman = 'sent'
                        share_log.save()
                        
                        results.append({
                            'nama': nama,
                            'nomor_wa': nomor_wa,
                            'success': True,
                            'share_url': share_url,
                            'message': '✓ Pesan berhasil dikirim ke WhatsApp'
                        })
                    else:
                        share_log.status_pengiriman = 'failed'
                        share_log.error_message = api_result.get('error', 'Unknown error')
                        share_log.save()
                        
                        results.append({
                            'nama': nama,
                            'nomor_wa': nomor_wa,
                            'success': False,
                            'share_url': share_url,
                            'message': f'✗ Gagal kirim: {api_result.get("message", "Unknown error")}'
                        })
                else:
                    # Just generate link, don't send via API (or API disabled)
                    wa_share_link = f"https://wa.me/{fontte._normalize_phone(nomor_wa) if fontte else nomor_wa.replace(' ', '')}?text={requests.utils.quote(message)}"
                    
                    results.append({
                        'nama': nama,
                        'nomor_wa': nomor_wa,
                        'success': True,
                        'share_url': share_url,
                        'message': '📱 Link checklist berhasil dibuat. Copy link di bawah atau kirim manual via WhatsApp.',
                        'wa_share_link': wa_share_link
                    })
            
            except Exception as e:
                logger.error(f"Error processing recipient {nama}: {str(e)}")
                results.append({
                    'nama': recipient.get('nama', 'Unknown'),
                    'nomor_wa': recipient.get('nomor_wa', ''),
                    'success': False,
                    'message': f'Error: {str(e)}'
                })
        
        # Build response
        response = {
            'success': True,
            'message': f'{sum(1 for r in results if r["success"])} dari {len(recipients)} penerima berhasil',
            'results': results,
            'is_first_submission': is_first_submission,
        }
        
        # Add warning jika ini edit mode
        if warning_message:
            response['warning'] = warning_message
        
        return JsonResponse(response)
    
    except Exception as e:
        logger.error(f"Error generate_share_link: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}'
        }, status=500)


@csrf_exempt  # Allow public access tanpa login
def checklist_fill_view(request, token):
    """
    Public view untuk isi checklist via share link.
    Tidak perlu login.
    
    GET: Show form checklist
    """
    from preventive_jobs.whatsapp_utils import verify_checklist_share_token
    from preventive_jobs.models import ChecklistShareLog
    
    try:
        # Verify token
        token_data = verify_checklist_share_token(token)
        
        if not token_data:
            return render(request, 'preventive_jobs/checklist_fill_error.html', {
                'error_title': 'Link Tidak Valid atau Expired',
                'error_message': 'Link untuk mengisi checklist sudah kadaluarsa atau tidak valid. Silakan minta link baru dari PIC.'
            }, status=400)
        
        execution_id = token_data.get('execution_id')
        checklist_result_id = token_data.get('checklist_result_id')
        
        # Get execution & checklist
        execution = get_object_or_404(PreventiveJobExecution, id=execution_id)
        checklist_result = get_object_or_404(ChecklistResult, id=checklist_result_id)
        
        # Get checklist template dengan fallback:
        # 1. Dari checklist_result (yang di-attach saat create)
        # 2. Fallback ke execution.template.checklist_template
        checklist_template = checklist_result.checklist_template
        if not checklist_template and execution.template:
            checklist_template = execution.template.checklist_template
        
        logger.info(f"checklist_fill_view execution_id={execution_id}, checklist_result_id={checklist_result_id}")
        logger.info(f"  checklist_result.checklist_template: {checklist_result.checklist_template}")
        logger.info(f"  execution.template.checklist_template: {execution.template.checklist_template if execution.template else 'N/A'}")
        logger.info(f"  Final checklist_template: {checklist_template}")
        
        if not checklist_template:
            logger.error(f"Checklist template tidak ditemukan untuk execution {execution_id}")
            logger.error(f"  checklist_result.checklist_template: {checklist_result.checklist_template}")
            logger.error(f"  execution.template: {execution.template}")
            return render(request, 'preventive_jobs/checklist_fill_error.html', {
                'error_title': 'Checklist Tidak Tersedia',
                'error_message': 'Checklist template tidak dikonfigurasi untuk job preventive ini.'
            }, status=400)
        
        # Track access & update status Pending -> Partial
        ChecklistShareLog.objects.filter(share_token=token).update(
            accessed_at=timezone.now()
        )
        
        # Update status dari Pending -> Partial saat form dibuka
        if checklist_result.status_overall == 'Pending':
            checklist_result.status_overall = 'Partial'
            checklist_result.save(update_fields=['status_overall'])
        
        # Get checklist items dari template yang di-attach di job
        checklist_items = checklist_template.items.all().order_by('no_urut')
        
        logger.info(f"Checklist items count: {checklist_items.count()}")
        for item in checklist_items:
            logger.info(f"  Item {item.no_urut}: {item.item_pemeriksaan} (Type: {item.item_type})")
        
        # Convert items to JSON untuk reliable rendering di JS
        items_json = []
        for item in checklist_items:
            items_json.append({
                'id': item.id,
                'no_urut': item.no_urut,
                'item_pemeriksaan': item.item_pemeriksaan,
                'item_type': item.item_type,
                'standar_normal': item.standar_normal or '',
                'unit': item.unit or '',
                'nilai_min': float(item.nilai_min) if item.nilai_min else None,
                'nilai_max': float(item.nilai_max) if item.nilai_max else None,
                'text_options': item.text_options or '',
            })
        
        # Pre-populate dengan hasil pengukuran yang sudah ada (jika ada)
        item_data = {}
        hasil_pengukuran = checklist_result.hasil_pengukuran or {}
        
        for item in checklist_items:
            item_id = str(item.id)
            if item_id in hasil_pengukuran:
                item_data[item_id] = hasil_pengukuran[item_id]
            else:
                item_data[item_id] = {
                    'nilai': '',
                    'status': '',
                    'remark': ''
                }
        
        logger.info(f"Item data prepared: {json.dumps(item_data, indent=2)}")
        
        context = {
            'execution': execution,
            'checklist_result': checklist_result,
            'checklist_template': checklist_template,
            'checklist_items': checklist_items,
            'checklist_items_json': json.dumps(items_json),  # ← JSON array dari items
            'item_data': json.dumps(item_data),
            'token': token,
        }
        
        logger.info(f"Context checklist_items: {[item.item_pemeriksaan for item in checklist_items]}")
        logger.info(f"Context checklist_items count: {checklist_items.count()}")
        logger.info(f"Items JSON: {json.dumps(items_json, indent=2)}")
        
        return render(request, 'preventive_jobs/checklist_fill_form.html', context)
    
    except PreventiveJobExecution.DoesNotExist:
        return render(request, 'preventive_jobs/checklist_fill_error.html', {
            'error_title': 'Execution Tidak Ditemukan',
            'error_message': 'Job execution tidak ditemukan.'
        }, status=404)
    
    except ChecklistResult.DoesNotExist:
        return render(request, 'preventive_jobs/checklist_fill_error.html', {
            'error_title': 'Checklist Tidak Ditemukan',
            'error_message': 'Checklist untuk job ini tidak ditemukan.'
        }, status=404)
    
    except Exception as e:
        logger.error(f"Error checklist_fill_view: {str(e)}")
        return render(request, 'preventive_jobs/checklist_fill_error.html', {
            'error_title': 'Terjadi Kesalahan',
            'error_message': f'Silakan hubungi administrator: {str(e)}'
        }, status=500)


@csrf_exempt  # Allow public submission
@require_http_methods(["POST"])
def save_checklist_via_token(request):
    """
    Public API untuk save checklist hasil dari share link.
    
    POST Data:
        - token: share token
        - nama_pengisi: nama orang yang isi
        - hasil_pengukuran: JSON mapping item_id -> {nilai, status, remark}
        - status_overall: OK / NG / Partial
        - catatan: catatan tambahan
    
    Returns:
        JSON response
    """
    from preventive_jobs.whatsapp_utils import (
        verify_checklist_share_token,
        send_checklist_filled_notification
    )
    from preventive_jobs.models import ChecklistShareLog
    
    try:
        token = request.POST.get('token', '').strip()
        
        if not token:
            return JsonResponse({
                'success': False,
                'message': 'Token tidak valid'
            }, status=400)
        
        # Verify token
        token_data = verify_checklist_share_token(token)
        
        if not token_data:
            return JsonResponse({
                'success': False,
                'message': 'Link sudah kadaluarsa'
            }, status=400)
        
        execution_id = token_data.get('execution_id')
        checklist_result_id = token_data.get('checklist_result_id')
        
        execution = get_object_or_404(PreventiveJobExecution, id=execution_id)
        checklist_result = get_object_or_404(ChecklistResult, id=checklist_result_id)
        
        # Parse hasil pengukuran
        hasil_pengukuran_json = request.POST.get('hasil_pengukuran', '{}')
        hasil_pengukuran = json.loads(hasil_pengukuran_json)
        
        status_overall = request.POST.get('status_overall', 'Partial').strip()
        catatan = request.POST.get('catatan', '').strip()
        nama_pengisi = request.POST.get('nama_pengisi', 'Unknown').strip()
        
        # Update checklist result
        checklist_result.hasil_pengukuran = hasil_pengukuran
        checklist_result.status_overall = status_overall
        checklist_result.catatan_umum = catatan
        checklist_result.accessed_by_name = nama_pengisi
        checklist_result.submitted_at = timezone.now()
        checklist_result.is_submitted_via_share = True
        checklist_result.tanggal_pengisian = timezone.now()
        checklist_result.save()
        
        # Update share log
        share_log = ChecklistShareLog.objects.filter(share_token=token).first()
        if share_log:
            share_log.submitted_at = timezone.now()
            share_log.save()
        
        # Send notification to PIC
        send_checklist_filled_notification(execution, checklist_result, share_log)
        
        # Update execution status if all checklist items are filled
        if status_overall in ['OK', 'NG']:
            execution.status = 'Done'
            execution.actual_date = timezone.now().date()
            execution.save()
        
        return JsonResponse({
            'success': True,
            'message': '✓ Checklist berhasil disimpan! Terima kasih.',
            'redirect_url': request.POST.get('redirect_url', '')
        })
    
    except PreventiveJobExecution.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Execution tidak ditemukan'
        }, status=404)
    
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Format data tidak valid'
        }, status=400)
    
    except Exception as e:
        logger.error(f"Error save_checklist_via_token: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}'
        }, status=500)

