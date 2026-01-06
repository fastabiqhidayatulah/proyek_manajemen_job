"""
Script untuk generate template Excel untuk import barang
Jalankan: python manage.py shell < create_import_template.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Import Barang"

# Define styles
header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)
example_fill = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
example_font = Font(size=10, italic=True, color="6B7280")

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = ["Nama Barang", "Kategori", "Spesifikasi", "Lokasi Penyimpanan", "Stok Awal"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# Row 1 height
ws.row_dimensions[1].height = 25

# Example data
examples = [
    ["Motor 3 Hp 2 Pole Foot", "mechanical", "Motor pisau, 3 Hp, 2 pole foot, 1500 RPM", "Rak A1", 5],
    ["Bearing 6203", "mechanical", "Bearing deep groove ball 6203, dimensi 17x40x12mm", "Rak B2", 10],
    ["Pompa Hidrolis P30", "hydraulic", "Pompa hidrolis gear pump P30, kapasitas 30cc/rev", "Gudang Utama", 2],
    ["O-ring Nitrile 25x3", "seals", "O-ring nitrile standar AS568, ukuran 25x3mm", "Rak C1", 50],
    ["Bolt M12x100 Grade 8.8", "fasteners", "Baut hex head M12x100, grade 8.8, SS304", "Rak D1", 100],
]

for row_num, example in enumerate(examples, 2):
    for col_num, value in enumerate(example, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.font = example_font
        cell.fill = example_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[row_num].height = 30

# Add instruction rows
instruction_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
instruction_font = Font(size=9, color="92400E")

start_row = len(examples) + 3
ws.merge_cells(f'A{start_row}:E{start_row}')
instr_cell = ws.cell(row=start_row, column=1)
instr_cell.value = "📝 PANDUAN PENGISIAN"
instr_cell.font = Font(bold=True, size=11, color="1E40AF")

# Instructions
instructions = [
    ("Nama Barang", "WAJIB - Nama atau tipe barang"),
    ("Kategori", "WAJIB - Pilih: electrical, mechanical, hydraulic, seals, fasteners, lubricants, consumables"),
    ("Spesifikasi", "OPSIONAL - Detail teknis barang"),
    ("Lokasi Penyimpanan", "OPSIONAL - Lokasi gudang atau rak penyimpanan"),
    ("Stok Awal", "OPSIONAL - Jumlah awal (angka saja, default: 0)"),
]

current_row = start_row + 2
for col_num, (field, instruction) in enumerate(instructions, 1):
    cell_field = ws.cell(row=current_row, column=1)
    cell_field.value = field
    cell_field.font = Font(bold=True, size=10)
    
    cell_instr = ws.cell(row=current_row, column=2)
    cell_instr.value = instruction
    ws.merge_cells(f'B{current_row}:E{current_row}')
    
    current_row += 1

# Column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 15

# Save file
output_path = "static/templates/Barang_Import_Template.xlsx"
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"✓ Template Excel berhasil dibuat: {output_path}")
