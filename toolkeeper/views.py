from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse_lazy
from django.utils import timezone
from django.db.models import Q
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse

from .models import Tool, Peminjaman, DetailPeminjaman, Pengembalian, DetailPengembalian
from .forms import (
    ToolForm, ToolImportForm, PeminjamanForm, DetailPeminjamanFormSet,
    DetailPeminjamanFormSetWithStockValidation,
    PengembalianForm, DetailPengembalianFormSet
)
from core.models import Karyawan


# ==============================================================================
# PERMISSION HELPER
# ==============================================================================
def can_manage_toolkeeper(user):
    """Check apakah user bisa manage tool keeper (Workshop atau Warehouse group)"""
    if not user.is_authenticated:
        return False
    return user.groups.filter(name__in=['Workshop', 'Warehouse']).exists()


# ==============================================================================
# TOOL VIEWS
# ==============================================================================
class ToolListView(LoginRequiredMixin, ListView):
    """List semua tools/alat"""
    model = Tool
    template_name = 'toolkeeper/tool_list.html'
    context_object_name = 'tools'
    paginate_by = 50
    login_url = 'login'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['can_manage'] = can_manage_toolkeeper(self.request.user)
        return context


class ToolCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """Create tool baru"""
    model = Tool
    form_class = ToolForm
    template_name = 'toolkeeper/tool_form.html'
    success_url = reverse_lazy('toolkeeper:tool-list')
    login_url = 'login'
    
    def test_func(self):
        return can_manage_toolkeeper(self.request.user)
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f"Tool '{self.object.nama}' berhasil dibuat")
        return response


class ToolEditView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Edit tool"""
    model = Tool
    form_class = ToolForm
    template_name = 'toolkeeper/tool_form.html'
    success_url = reverse_lazy('toolkeeper:tool-list')
    login_url = 'login'
    
    def test_func(self):
        return can_manage_toolkeeper(self.request.user)
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f"Tool '{self.object.nama}' berhasil diupdate")
        return response


class ToolDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """Delete tool"""
    model = Tool
    template_name = 'toolkeeper/tool_confirm_delete.html'
    success_url = reverse_lazy('toolkeeper:tool-list')
    login_url = 'login'
    
    def test_func(self):
        return can_manage_toolkeeper(self.request.user)
    
    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        tool_nama = self.object.nama
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f"Tool '{tool_nama}' berhasil dihapus")
        return response


@login_required(login_url='core:login')
def tool_import_view(request):
    """Import tools dari Excel"""
    if not can_manage_toolkeeper(request.user):
        messages.error(request, "Akses ditolak")
        return redirect('toolkeeper:tool-list')
    
    if request.method == 'POST':
        form = ToolImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            try:
                import openpyxl
                from openpyxl.utils import get_column_letter
                
                # Load workbook
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                
                success_count = 0
                error_count = 0
                error_rows = []
                
                # Skip header (row 1)
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                    try:
                        # Get cell values
                        nama_cell = row[0]
                        spesifikasi_cell = row[1]
                        jumlah_cell = row[2]
                        
                        if not nama_cell.value:
                            error_rows.append(f"Row {row_idx}: Nama alat kosong")
                            error_count += 1
                            continue
                        
                        nama = str(nama_cell.value).strip()
                        spesifikasi = str(spesifikasi_cell.value) if spesifikasi_cell.value else ""
                        
                        try:
                            jumlah = int(jumlah_cell.value) if jumlah_cell.value else 0
                        except (ValueError, TypeError):
                            error_rows.append(f"Row {row_idx}: Jumlah harus angka (nilai: {jumlah_cell.value})")
                            error_count += 1
                            continue
                        
                        # Check if already exists
                        if Tool.objects.filter(nama=nama).exists():
                            error_rows.append(f"Row {row_idx}: '{nama}' sudah ada di database")
                            error_count += 1
                            continue
                        
                        # Create tool
                        Tool.objects.create(
                            nama=nama,
                            spesifikasi=spesifikasi,
                            jumlah_total=jumlah
                        )
                        success_count += 1
                        
                    except Exception as e:
                        error_rows.append(f"Row {row_idx}: {str(e)}")
                        error_count += 1
                
                # Show results
                if success_count > 0:
                    messages.success(request, f"✓ Berhasil import {success_count} tool!")
                
                if error_count > 0:
                    error_msg = f"⚠ Ada {error_count} baris yang gagal:\n" + "\n".join(error_rows[:5])
                    if len(error_rows) > 5:
                        error_msg += f"\n... dan {len(error_rows)-5} error lainnya"
                    messages.warning(request, error_msg)
                
                return redirect('toolkeeper:tool-list')
                
            except Exception as e:
                messages.error(request, f"Error membaca file Excel: {str(e)}")
    else:
        form = ToolImportForm()
    
    context = {
        'form': form,
        'can_manage': can_manage_toolkeeper(request.user),
    }
    return render(request, 'toolkeeper/tool_import.html', context)


# ==============================================================================
# PEMINJAMAN VIEWS
# ==============================================================================
class PeminjamanListView(LoginRequiredMixin, ListView):
    """List peminjaman aktif & overdue"""
    model = Peminjaman
    template_name = 'toolkeeper/peminjaman_list.html'
    context_object_name = 'peminjaman_list'
    paginate_by = 50
    login_url = 'login'
    
    def get_queryset(self):
        queryset = Peminjaman.objects.select_related('peminjam').prefetch_related('detail_peminjaman')
        
        # Check and update status untuk semua peminjaman aktif/overdue
        # 1. Check jika deadline lewat tapi masih aktif → mark sebagai overdue
        # 2. Check jika semua alat kembali → mark sebagai selesai
        from django.utils import timezone
        now = timezone.now()
        
        for peminjaman in queryset.filter(status__in=['aktif', 'overdue']):
            # Update ke overdue jika tgl_rencana_kembali sudah lewat
            if peminjaman.status == 'aktif' and now > peminjaman.tgl_rencana_kembali:
                peminjaman.status = 'overdue'
                peminjaman.save(update_fields=['status'])
            
            # Update ke selesai jika semua alat sudah dikembalikan
            peminjaman.check_and_update_status()
        
        # Filter by status
        status = self.request.GET.get('status', '')
        if status:
            queryset = queryset.filter(status=status)
        else:
            # Default: tampilkan aktif & overdue
            queryset = queryset.filter(status__in=['aktif', 'overdue'])
        
        # Search by peminjam
        search = self.request.GET.get('search', '')
        if search:
            queryset = queryset.filter(peminjam__nama_lengkap__icontains=search)
        
        return queryset.order_by('-tgl_pinjam')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['can_manage'] = can_manage_toolkeeper(self.request.user)
        context['status'] = self.request.GET.get('status', '')
        context['search'] = self.request.GET.get('search', '')
        context['view_type'] = 'peminjam'  # Show this is peminjam view
        
        # Get per-alat data for context
        detail_peminjamannya = DetailPeminjaman.objects.select_related(
            'peminjaman__peminjam', 'tool'
        ).prefetch_related('peminjaman__pengembalian__detail_pengembalian')
        
        # Group by tool
        tools_data = {}
        for detail in detail_peminjamannya:
            tool_id = detail.tool.id
            if tool_id not in tools_data:
                tools_data[tool_id] = {
                    'tool': detail.tool,
                    'peminjaman': []
                }
            tools_data[tool_id]['peminjaman'].append(detail)
        
        context['tools_data'] = tools_data.values()
        
        return context


class PeminjamanCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """Create peminjaman baru"""
    model = Peminjaman
    form_class = PeminjamanForm
    template_name = 'toolkeeper/peminjaman_form.html'
    login_url = 'login'
    
    def test_func(self):
        return can_manage_toolkeeper(self.request.user)
    
    def get(self, request, *args, **kwargs):
        response = super().get(request, *args, **kwargs)
        response.context_data['formset'] = DetailPeminjamanFormSetWithStockValidation()
        response.context_data['karyawan_list'] = Karyawan.objects.filter(status='Aktif').order_by('nama_lengkap')
        response.context_data['tool_list'] = Tool.objects.all().order_by('nama')
        return response
    
    def post(self, request, *args, **kwargs):
        self.object = None
        form = self.get_form()
        formset = DetailPeminjamanFormSetWithStockValidation(self.request.POST)
        
        if form.is_valid() and formset.is_valid():
            self.object = form.save(commit=False)
            self.object.created_by = request.user
            self.object.save()
            
            formset.instance = self.object
            formset.save()
            
            messages.success(request, "Peminjaman berhasil dibuat")
            return redirect(self.get_success_url())
        else:
            return self.form_invalid(form)
    
    def get_success_url(self):
        return reverse_lazy('toolkeeper:peminjaman-detail', kwargs={'pk': self.object.pk})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = DetailPeminjamanFormSet(self.request.POST)
        else:
            context['formset'] = DetailPeminjamanFormSet()
        context['karyawan_list'] = Karyawan.objects.filter(status='Aktif').order_by('nama_lengkap')
        context['tool_list'] = Tool.objects.all().order_by('nama')
        return context


class PeminjamanDetailView(LoginRequiredMixin, DetailView):
    """Detail peminjaman"""
    model = Peminjaman
    template_name = 'toolkeeper/peminjaman_detail.html'
    context_object_name = 'peminjaman'
    login_url = 'login'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['can_manage'] = can_manage_toolkeeper(self.request.user)
        context['pengembalian_list'] = self.object.pengembalian.all()
        return context


class PeminjamanEditView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Edit peminjaman"""
    model = Peminjaman
    form_class = PeminjamanForm
    template_name = 'toolkeeper/peminjaman_form.html'
    login_url = 'login'
    
    def test_func(self):
        return can_manage_toolkeeper(self.request.user) and self.get_object().status in ['aktif', 'overdue']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = DetailPeminjamanFormSet(self.request.POST, instance=self.object)
        else:
            context['formset'] = DetailPeminjamanFormSet(instance=self.object)
        context['karyawan_list'] = Karyawan.objects.filter(status='Aktif').order_by('nama_lengkap')
        context['tool_list'] = Tool.objects.all().order_by('nama')
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        
        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()
            messages.success(self.request, "Peminjaman berhasil diupdate")
            return redirect(self.get_success_url())
        else:
            return self.form_invalid(form)
    
    def get_success_url(self):
        return reverse_lazy('toolkeeper:peminjaman-detail', kwargs={'pk': self.object.pk})


# ==============================================================================
# PENGEMBALIAN VIEWS
# ==============================================================================
class PengembalianCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """Create pengembalian (return tools)"""
    model = Pengembalian
    form_class = PengembalianForm
    template_name = 'toolkeeper/pengembalian_form.html'
    login_url = 'login'
    
    def test_func(self):
        peminjaman = get_object_or_404(Peminjaman, pk=self.kwargs['pk'])
        return can_manage_toolkeeper(self.request.user) and peminjaman.status in ['aktif', 'overdue']
    
    def get_object(self):
        return get_object_or_404(Peminjaman, pk=self.kwargs['pk'])
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        peminjaman = self.get_object()
        # Pre-fill dengan peminjam
        if self.request.method == 'GET':
            kwargs['initial'] = {'dikembalikan_oleh': peminjaman.peminjam}
        return kwargs
    
    def get(self, request, *args, **kwargs):
        peminjaman = self.get_object()
        response = super().get(request, *args, **kwargs)
        
        # Pre-populate formset dengan detail peminjaman yang belum dikembalikan lengkap
        if 'formset' not in response.context_data:
            initial_data = []
            for detail in peminjaman.detail_peminjaman.all():
                if detail.qty_belum_kembali > 0:
                    initial_data.append({
                        'tool': detail.tool,
                        'qty_kembali': detail.qty_belum_kembali,
                    })
            response.context_data['formset'] = DetailPengembalianFormSet(initial=initial_data)
        
        response.context_data['peminjaman'] = peminjaman
        response.context_data['karyawan_list'] = Karyawan.objects.filter(status='Aktif').order_by('nama_lengkap')
        response.context_data['tool_list'] = Tool.objects.all().order_by('nama')
        return response
    
    def post(self, request, *args, **kwargs):
        peminjaman = self.get_object()
        self.object = None
        form = self.get_form()
        formset = DetailPengembalianFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            self.object = form.save(commit=False)
            self.object.peminjaman = peminjaman
            self.object.save()
            
            formset.instance = self.object
            formset.save()
            
            # Check dan auto-update status ke selesai jika semua alat sudah dikembalikan
            peminjaman.check_and_update_status()
            
            messages.success(request, "Pengembalian berhasil dicatat")
            return redirect(reverse_lazy('toolkeeper:peminjaman-detail', kwargs={'pk': peminjaman.pk}))
        else:
            context = self.get_context_data()
            context['formset'] = formset
            context['peminjaman'] = peminjaman
            return self.render_to_response(context)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['formset'] = DetailPengembalianFormSet(self.request.POST)
        context['peminjaman'] = self.get_object()
        return context


class PengembalianDetailView(LoginRequiredMixin, DetailView):
    """Detail pengembalian"""
    model = Pengembalian
    template_name = 'toolkeeper/pengembalian_detail.html'
    context_object_name = 'pengembalian'
    login_url = 'login'


# ==============================================================================
# REPORT VIEW
# ==============================================================================
class ReportView(LoginRequiredMixin, ListView):
    """Report peminjaman & pengembalian"""
    template_name = 'toolkeeper/report.html'
    login_url = 'login'
    
    def get_context_data(self, **kwargs):
        context = {}
        
        # Summary stats
        context['total_peminjaman_aktif'] = Peminjaman.objects.filter(status='aktif').count()
        context['total_peminjaman_overdue'] = Peminjaman.objects.filter(status='overdue').count()
        context['total_peminjaman_selesai'] = Peminjaman.objects.filter(status='selesai').count()
        
        # Tools
        context['tools_list'] = Tool.objects.all()
        
        # Overdue peminjaman
        context['overdue_peminjaman'] = Peminjaman.objects.filter(status='overdue').select_related('peminjam')
        
        # Peminjaman terakhir
        context['recent_peminjaman'] = Peminjaman.objects.select_related('peminjam')[:10]
        
        return context


# ==============================================================================
# API ENDPOINTS
# ==============================================================================
@login_required
def api_return_tool(request):
    """API endpoint untuk return per alat (partial return)"""
    import json
    from django.http import JsonResponse
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    if not can_manage_toolkeeper(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        
        peminjaman_id = data.get('peminjaman_id')
        detail_id = data.get('detail_id')
        qty_kembali = int(data.get('qty_kembali', 0))
        kondisi_kembali = data.get('kondisi_kembali')
        catatan = data.get('catatan', '')
        
        # Validation
        if not all([peminjaman_id, detail_id, qty_kembali, kondisi_kembali]):
            return JsonResponse({'success': False, 'error': 'Missing required fields'}, status=400)
        
        # Get objects
        peminjaman = get_object_or_404(Peminjaman, id=peminjaman_id)
        detail = get_object_or_404(DetailPeminjaman, id=detail_id, peminjaman=peminjaman)
        
        # Validate qty
        if qty_kembali > detail.qty_belum_kembali:
            return JsonResponse({'success': False, 'error': 'Quantity exceeds remaining amount'}, status=400)
        
        # Create Pengembalian if doesn't exist for today
        pengembalian, created = Pengembalian.objects.get_or_create(
            peminjaman=peminjaman,
            tgl_kembali__date=timezone.now().date(),
            defaults={
                'tgl_kembali': timezone.now(),
                'dikembalikan_oleh': request.user.karyawan if hasattr(request.user, 'karyawan') else peminjaman.peminjam,
                'catatan': catatan
            }
        )
        
        # Create detail pengembalian
        detail_return, created = DetailPengembalian.objects.get_or_create(
            pengembalian=pengembalian,
            tool=detail.tool,
            defaults={
                'qty_kembali': qty_kembali,
                'kondisi_kembali': kondisi_kembali
            }
        )
        
        # If exists, update qty
        if not created:
            detail_return.qty_kembali += qty_kembali
            detail_return.kondisi_kembali = kondisi_kembali
            detail_return.save()
        
        # Check if peminjaman complete (qty_kembali calculated from DetailPengembalian)
        peminjaman.check_and_update_status()
        
        return JsonResponse({'success': True, 'message': 'Return recorded successfully'})
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(['GET'])
def api_get_tool_stock(request, tool_id):
    """API endpoint untuk get stok tool"""
    try:
        tool = Tool.objects.get(id=tool_id)
        
        return JsonResponse({
            'success': True,
            'tool_id': str(tool.id),
            'tool_name': tool.nama,
            'jumlah_total': tool.jumlah_total,
            'jumlah_tersedia': tool.jumlah_tersedia,
            'message': f'{tool.nama}: {tool.jumlah_tersedia} / {tool.jumlah_total} tersedia'
        })
    except Tool.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Tool tidak ditemukan'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(['POST'])
def api_add_karyawan(request):
    """API endpoint untuk add karyawan baru (quick add dari form peminjaman)"""
    import json
    
    if not can_manage_toolkeeper(request.user):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        
        nik = data.get('nik', '').strip()
        nama_lengkap = data.get('nama_lengkap', '').strip()
        
        # Validation
        if not nik or not nama_lengkap:
            return JsonResponse({'success': False, 'error': 'NIK dan Nama Lengkap harus diisi'}, status=400)
        
        # Check if NIK already exists
        if Karyawan.objects.filter(nik=nik).exists():
            return JsonResponse({'success': False, 'error': f'NIK {nik} sudah terdaftar'}, status=400)
        
        # Create karyawan (only required fields)
        karyawan = Karyawan.objects.create(
            nik=nik,
            nama_lengkap=nama_lengkap,
            status='Aktif'  # Default status aktif
        )
        
        return JsonResponse({
            'success': True, 
            'karyawan_id': str(karyawan.id),
            'karyawan_name': f'{karyawan.nik} - {karyawan.nama_lengkap}'
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
